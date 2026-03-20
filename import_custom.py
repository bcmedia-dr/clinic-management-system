from openpyxl import load_workbook
import re
from phone_utils import format_phone, normalize_specialty
from sqlalchemy.exc import IntegrityError


def _normalize_phone(phone):
    """移除所有非數字字元，用於重複比對"""
    return re.sub(r'\D', '', str(phone)) if phone else ''


def _parse_name(raw_name):
    """
    清理院名，提取備註內容。
    - 全形/半形括號內文字 → note
    - 斜線後文字 → note（與括號備註以「 / 」合併）
    返回 (name, note)
    """
    if not raw_name:
        return '', ''

    raw_name = str(raw_name).strip()
    notes = []

    # 提取括號內容（支援全形 （）和半形 ()）
    bracket_pattern = re.compile(r'[（(]([^）)]*)[）)]')
    for m in bracket_pattern.findall(raw_name):
        if m.strip():
            notes.append(m.strip())
    name = bracket_pattern.sub('', raw_name).strip()

    # 提取斜線後文字
    if '/' in name:
        parts = name.split('/', 1)
        name = parts[0].strip()
        slash_part = parts[1].strip()
        if slash_part:
            notes.append(slash_part)

    note = ' / '.join(notes)
    return name.strip(), note


def import_custom_clinics(file_path, db, Clinic):
    """
    匯入自訂格式 Excel（欄位名稱：縣市、區域、院名、科別、院址、電話、聯絡人、院長）。
    院名自動清理括號/斜線內容為備註。
    以電話號碼檢查重複，重複者跳過。
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        header_row = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        col = {name: idx for idx, name in enumerate(header_row)}

        if '院名' not in col:
            return {'success': False, 'error': 'Excel 缺少必要欄位：院名'}

        def _get(row, field):
            idx = col.get(field)
            if idx is None or idx >= len(row):
                return ''
            v = row[idx]
            return str(v).strip() if v is not None else ''

        # 建立現有電話 → 診所物件對應表（用 phone_normalized 為鍵，方便查詢與更新）
        phone_to_existing = {}
        for c in Clinic.query.all():
            if c.phone_normalized:
                phone_to_existing[c.phone_normalized] = c

        imported_count = 0
        updated_count  = 0
        error_count    = 0
        error_details  = []

        consecutive_blank = 0  # 連續空白行計數器
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row):
                    consecutive_blank += 1
                    if consecutive_blank >= 5:  # 連續5筆空白即停止
                        break
                    continue
                consecutive_blank = 0  # 有資料就重置

                raw_name = _get(row, '院名')
                if not raw_name:
                    error_count += 1
                    error_details.append(f'第{row_num}列：院名為必填')
                    continue

                name, extracted_note = _parse_name(raw_name)
                if not name:
                    error_count += 1
                    error_details.append(f'第{row_num}列：清理後院名為空（原始值：{raw_name}）')
                    continue

                phone_raw  = _get(row, '電話')
                phone_fmt  = format_phone(phone_raw, _get(row, '縣市') or None)   # 先補區碼
                normalized = _normalize_phone(phone_fmt)                            # 再正規化（含區碼的完整數字）

                if normalized and normalized in phone_to_existing:
                    # 電話已存在 → 更新現有診所資料（phone / phone_normalized 不動）
                    existing = phone_to_existing[normalized]
                    existing.region         = _get(row, '縣市') or None
                    existing.district       = _get(row, '區域') or None
                    existing.name           = name
                    existing.specialties    = normalize_specialty(_get(row, '科別')) or None
                    existing.address        = _get(row, '院址') or None
                    existing.contact_person = _get(row, '聯絡人') or None
                    # 備註：有新值才覆蓋，保留舊備註
                    if extracted_note:
                        existing.note = extracted_note
                    updated_count += 1
                    continue

                clinic = Clinic(
                    region           = _get(row, '縣市') or None,
                    district         = _get(row, '區域') or None,
                    name             = name,
                    specialties      = normalize_specialty(_get(row, '科別')) or None,
                    address          = _get(row, '院址') or None,
                    phone            = phone_fmt or None,
                    phone_normalized = normalized or None,  # format_phone 後的完整數字
                    contact_person   = _get(row, '聯絡人') or None,
                    note             = extracted_note or None,
                )
                try:
                    # 用 savepoint 隔離每筆 INSERT，UniqueViolation 只回滾該筆
                    sp = db.session.begin_nested()
                    db.session.add(clinic)
                    db.session.flush()
                    if normalized:
                        phone_to_existing[normalized] = clinic
                    imported_count += 1
                except IntegrityError:
                    # 並發情況：INSERT 失敗但 phone_to_existing 未命中，改為更新
                    sp.rollback()
                    existing = Clinic.query.filter_by(phone_normalized=normalized).first()
                    if existing:
                        existing.region         = _get(row, '縣市') or None
                        existing.district       = _get(row, '區域') or None
                        existing.name           = name
                        existing.specialties    = normalize_specialty(_get(row, '科別')) or None
                        existing.address        = _get(row, '院址') or None
                        existing.contact_person = _get(row, '聯絡人') or None
                        if extracted_note:
                            existing.note = extracted_note
                        phone_to_existing[normalized] = existing
                        updated_count += 1
                    else:
                        error_count += 1
                        error_details.append(f'第{row_num}列：電話衝突但查無既有診所，跳過')

            except Exception as e:
                error_count += 1
                error_details.append(f'第{row_num}列：{str(e)}')

        db.session.commit()

        return {
            'success':      True,
            'imported':     imported_count,
            'updated':      updated_count,
            'errors':       error_count,
            'error_details': error_details,
        }

    except Exception as e:
        db.session.rollback()
        return {'success': False, 'error': str(e)}
