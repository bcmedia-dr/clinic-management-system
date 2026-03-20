from openpyxl import load_workbook
from datetime import datetime
import re
from phone_utils import format_phone, normalize_specialty
from sqlalchemy.exc import IntegrityError


def _normalize_phone(phone):
    return re.sub(r'\D', '', str(phone)) if phone else ''


def import_clinics(file_path, db, Clinic, dry_run=False):
    """從 Excel 匯入診所資料，以電話號碼檢查重複。
    dry_run=True 時執行所有判斷但不寫入資料庫，回傳預覽結果。"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        expected_headers = ['縣市', '區域', '診所名稱', '科別', '地址', '電話', '負責人']
        actual_headers = [cell.value for cell in ws[1]]

        if actual_headers[:7] != expected_headers:
            return {'success': False, 'error': '檔案格式錯誤：標題列不符合要求'}

        # 建立現有電話 → 診所物件對應表（用 phone_normalized 為鍵，方便查詢與更新）
        phone_to_existing = {}
        for c in Clinic.query.all():
            if c.phone_normalized:
                phone_to_existing[c.phone_normalized] = c

        imported_count = 0
        updated_count  = 0
        errors = []
        preview = []  # 預覽清單（最多 10 筆）

        consecutive_blank = 0  # 連續空白行計數器
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row[:7]):
                    consecutive_blank += 1
                    if consecutive_blank >= 5:  # 連續5筆空白即停止，避免掃描百萬空白行
                        break
                    continue
                consecutive_blank = 0  # 有資料就重置

                region         = row[0]
                district       = row[1]
                name           = row[2]
                specialties    = row[3]
                address        = row[4]
                phone          = row[5]
                contact_person = row[6]

                if not name:
                    errors.append(f'第{row_num}列：診所名稱為必填')
                    continue

                phone_str  = str(phone) if phone else ''
                region_str = str(region).strip() if region else None
                phone_fmt  = format_phone(phone_str, region_str)   # 先補區碼
                normalized = _normalize_phone(phone_fmt)            # 再正規化（含區碼的完整數字）

                if normalized and normalized in phone_to_existing:
                    # 電話已存在 → 更新現有診所資料（phone / phone_normalized 不動）
                    existing = phone_to_existing[normalized]
                    existing.region         = region
                    existing.district       = district
                    existing.name           = name
                    existing.specialties    = normalize_specialty(str(specialties) if specialties else '') or None
                    existing.address        = address
                    existing.contact_person = contact_person
                    updated_count += 1
                    if len(preview) < 10:
                        preview.append({'name': str(name), 'phone': phone_fmt, 'action': 'update', 'region': region_str or ''})
                    continue

                clinic = Clinic(
                    region           = region,
                    district         = district,
                    name             = name,
                    specialties      = normalize_specialty(str(specialties) if specialties else '') or None,
                    address          = address,
                    phone            = phone_fmt,
                    phone_normalized = normalized or None,  # format_phone 後的完整數字
                    contact_person   = contact_person,
                )
                try:
                    # 用 savepoint 隔離每筆 INSERT，UniqueViolation 只回滾該筆
                    sp = db.session.begin_nested()
                    db.session.add(clinic)
                    db.session.flush()
                    if normalized:
                        phone_to_existing[normalized] = clinic
                    imported_count += 1
                    if len(preview) < 10:
                        preview.append({'name': str(name), 'phone': phone_fmt, 'action': 'create', 'region': region_str or ''})
                except IntegrityError:
                    # 並發情況：INSERT 失敗但 phone_to_existing 未命中，改為更新
                    sp.rollback()
                    existing = Clinic.query.filter_by(phone_normalized=normalized).first()
                    if existing:
                        existing.region         = region
                        existing.district       = district
                        existing.name           = name
                        existing.specialties    = normalize_specialty(str(specialties) if specialties else '') or None
                        existing.address        = address
                        existing.contact_person = contact_person
                        phone_to_existing[normalized] = existing
                        updated_count += 1
                        if len(preview) < 10:
                            preview.append({'name': str(name), 'phone': phone_fmt, 'action': 'update', 'region': region_str or ''})
                    else:
                        errors.append(f'第{row_num}列：電話衝突但查無既有診所，跳過')

            except Exception as e:
                errors.append(f'第{row_num}列：{str(e)}')

        # dry_run 模式：rollback 確保不寫入，回傳預覽結果
        if dry_run:
            db.session.rollback()
            return {
                'dry_run':      True,
                'would_create': imported_count,
                'would_update': updated_count,
                'would_skip':   0,
                'errors':       errors,
                'preview':      preview,
            }

        db.session.commit()

        return {
            'success':  True,
            'imported': imported_count,
            'updated':  updated_count,
            'errors':   errors,
        }

    except Exception as e:
        db.session.rollback()
        return {'success': False, 'error': str(e)}


def import_health_mall(file_path, db, HealthMall, dry_run=False):
    """從 Excel 匯入健康醫購資料，以電話號碼檢查重複。
    dry_run=True 時執行所有判斷但不寫入資料庫，回傳預覽結果。"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        expected_headers = ['縣市', '區域', '診所名稱', '科別', '地址', '電話', '負責人']
        actual_headers = [cell.value for cell in ws[1]]

        if actual_headers[:7] != expected_headers:
            return {'success': False, 'error': '檔案格式錯誤：標題列不符合要求'}

        header = [str(c.value).strip() if c.value else '' for c in ws[1]]
        col = {name: idx for idx, name in enumerate(header)}

        # 建立現有電話號碼集合
        existing_phones = set()
        for (phone,) in HealthMall.query.with_entities(HealthMall.phone).all():
            np = _normalize_phone(phone)
            if np:
                existing_phones.add(np)

        imported_count = 0
        skipped_count  = 0
        skipped_names  = []
        errors         = []
        preview        = []  # 預覽清單（最多 10 筆）

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row[:7]):
                    continue

                def _get(field):
                    idx = col.get(field)
                    if idx is None or idx >= len(row):
                        return ''
                    return str(row[idx]).strip() if row[idx] is not None else ''

                name = _get('診所名稱')
                if not name:
                    errors.append(f'第{row_num}列：診所名稱為必填')
                    continue

                phone_str  = _get('電話')
                normalized = _normalize_phone(phone_str)
                phone_fmt  = format_phone(phone_str, _get('縣市') or None)

                if normalized and normalized in existing_phones:
                    skipped_count += 1
                    skipped_names.append(name)
                    if len(preview) < 10:
                        preview.append({'name': name, 'phone': phone_fmt, 'action': 'skip', 'region': _get('縣市')})
                    continue

                start_date = None
                start_date_str = _get('開始日期')
                if start_date_str:
                    try:
                        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        pass

                status = _get('合作狀態') or '合作中'

                hm = HealthMall(
                    region         = _get('縣市'),
                    district       = _get('區域'),
                    name           = name,
                    specialties    = _get('科別'),
                    address        = _get('地址'),
                    phone          = phone_fmt,
                    contact_person = _get('負責人'),
                    status         = status,
                    start_date     = start_date,
                    note           = _get('備註'),
                )
                db.session.add(hm)
                if normalized:
                    existing_phones.add(normalized)
                imported_count += 1
                if len(preview) < 10:
                    preview.append({'name': name, 'phone': phone_fmt, 'action': 'create', 'region': _get('縣市')})

            except Exception as e:
                errors.append(f'第{row_num}列：{str(e)}')

        # dry_run 模式：rollback 確保不寫入，回傳預覽結果
        if dry_run:
            db.session.rollback()
            return {
                'dry_run':      True,
                'would_create': imported_count,
                'would_update': 0,
                'would_skip':   skipped_count,
                'errors':       errors,
                'preview':      preview,
            }

        db.session.commit()

        return {
            'success':       True,
            'imported':      imported_count,
            'skipped':       skipped_count,
            'skipped_names': skipped_names,
            'errors':        errors,
        }

    except Exception as e:
        db.session.rollback()
        return {'success': False, 'error': str(e)}
