from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text, func, case
from sqlalchemy.orm import joinedload
from sqlalchemy.exc import IntegrityError
from datetime import datetime, date
import os, re, math, hmac
from collections import Counter
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.utils import secure_filename
from export import export_clinics
from import_data import import_clinics, import_health_mall
from import_custom import import_custom_clinics
from phone_utils import format_phone, normalize_specialty, STANDARD_SPECIALTIES

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'fallback-only-for-local')

# 資料庫設定
DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:///clinics.db')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://')

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB 上傳限制

db = SQLAlchemy(app)


@app.errorhandler(413)
def file_too_large(e):
    return jsonify({'success': False, 'error': '檔案過大，最大允許 10MB'}), 413


# ── 資料模型 ─────────────────────────────────────────────────

class Clinic(db.Model):
    id               = db.Column(db.Integer, primary_key=True)
    region           = db.Column(db.String(50))
    district         = db.Column(db.String(50))
    name             = db.Column(db.String(200))
    specialties      = db.Column(db.String(500))
    address          = db.Column(db.String(300))
    phone            = db.Column(db.String(50))
    phone_normalized = db.Column(db.String(50), unique=True)        # 正規化電話（僅數字），用於唯一性約束
    contact_person   = db.Column(db.String(100))
    business_hours   = db.Column(db.String(200))
    note             = db.Column(db.Text)
    col_yaodai       = db.Column(db.Boolean, default=False)
    col_haibao       = db.Column(db.Boolean, default=False)
    col_paiyang      = db.Column(db.Boolean, default=False)
    col_baiwei       = db.Column(db.Boolean, default=False)
    status           = db.Column(db.String(20), default='active')   # 'active' | 'deleted'（軟刪除）
    deleted_at       = db.Column(db.DateTime)                       # 軟刪除時間（status='deleted' 時才有值）
    created_at       = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at       = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class HealthMall(db.Model):
    __tablename__ = 'health_mall'
    id             = db.Column(db.Integer, primary_key=True)
    clinic_id      = db.Column(db.Integer, db.ForeignKey('clinic.id'), nullable=True)
    region         = db.Column(db.String(50))
    district       = db.Column(db.String(50))
    name           = db.Column(db.String(200))
    phone          = db.Column(db.String(50))
    contact_person = db.Column(db.String(100))
    specialties    = db.Column(db.String(500))
    address        = db.Column(db.String(300))
    status         = db.Column(db.String(20), default='合作中')
    start_date     = db.Column(db.Date)
    note           = db.Column(db.Text)
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at     = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    clinic = db.relationship('Clinic', backref=db.backref('health_mall_records', lazy=True))


class Campaign(db.Model):
    __tablename__ = 'campaign'
    id                 = db.Column(db.Integer, primary_key=True)
    name               = db.Column(db.String(200), nullable=False)
    brand              = db.Column(db.String(100))
    year               = db.Column(db.Integer)
    month              = db.Column(db.Integer)
    note               = db.Column(db.Text)
    cooperation_items  = db.Column(db.String(200))  # 勾選的合作項目，逗號分隔（如 "海報,立牌,派樣"）
    cooperation_other  = db.Column(db.String(200))  # 其他合作項目，自由填寫
    created_at         = db.Column(db.DateTime, default=datetime.utcnow)


class CampaignClinic(db.Model):
    __tablename__ = 'campaign_clinic'
    id          = db.Column(db.Integer, primary_key=True)
    campaign_id = db.Column(db.Integer, db.ForeignKey('campaign.id'), nullable=False)
    clinic_id   = db.Column(db.Integer, db.ForeignKey('clinic.id'), nullable=False)
    joined_at   = db.Column(db.DateTime, default=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint('campaign_id', 'clinic_id', name='uq_campaign_clinic'),)

    campaign = db.relationship('Campaign', backref=db.backref('clinic_links', lazy=True))
    clinic   = db.relationship('Clinic',   backref=db.backref('campaign_links', lazy=True))


class BaiweiDoctor(db.Model):
    __tablename__ = 'baiwei_doctor'
    id          = db.Column(db.Integer, primary_key=True)
    clinic_id   = db.Column(db.Integer, db.ForeignKey('clinic.id', ondelete='SET NULL'), nullable=True)
    clinic_name = db.Column(db.String(200))
    region      = db.Column(db.String(50))
    district    = db.Column(db.String(50))
    address     = db.Column(db.String(300))
    phone       = db.Column(db.String(50))
    doctor_name = db.Column(db.String(100))
    specialty   = db.Column(db.String(200))
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)

    clinic = db.relationship('Clinic', backref=db.backref('baiwei_doctors', lazy=True))


class BaiweiCampaign(db.Model):
    __tablename__ = 'baiwei_campaign'
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(200), nullable=False)
    year       = db.Column(db.Integer)
    month      = db.Column(db.Integer)
    note       = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class BaiweiParticipation(db.Model):
    __tablename__ = 'baiwei_participation'
    id          = db.Column(db.Integer, primary_key=True)
    doctor_id   = db.Column(db.Integer, db.ForeignKey('baiwei_doctor.id', ondelete='CASCADE'), nullable=False)
    campaign_id = db.Column(db.Integer, db.ForeignKey('baiwei_campaign.id', ondelete='CASCADE'), nullable=False)
    __table_args__ = (db.UniqueConstraint('doctor_id', 'campaign_id', name='uq_baiwei_participation'),)

    doctor   = db.relationship('BaiweiDoctor',   backref=db.backref('participations', lazy=True))
    campaign = db.relationship('BaiweiCampaign', backref=db.backref('participations', lazy=True))


class MatchHistory(db.Model):
    """活動比對歷史：儲存每次比對的結果摘要"""
    __tablename__ = 'match_history'
    id             = db.Column(db.Integer, primary_key=True)
    campaign_name  = db.Column(db.String(200))
    matched_count  = db.Column(db.Integer, default=0)
    not_joined_count   = db.Column(db.Integer, default=0)
    not_in_system_count = db.Column(db.Integer, default=0)
    result_json    = db.Column(db.Text)    # 完整比對結果 JSON
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)


class AuditLog(db.Model):
    """操作記錄：記錄診所的新增、編輯、刪除等操作"""
    __tablename__ = 'audit_log'
    id          = db.Column(db.Integer, primary_key=True)
    action      = db.Column(db.String(20))    # '新增' / '編輯' / '刪除' / '永久刪除'
    target_name = db.Column(db.String(200))   # 操作的診所名稱
    target_id   = db.Column(db.Integer)       # 操作的診所 ID
    detail      = db.Column(db.Text)          # 變更摘要（可空）
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)


def _write_audit(action, name, target_id, detail=None):
    """寫入一筆操作記錄（呼叫端需自行 commit）"""
    log = AuditLog(action=action, target_name=name, target_id=target_id, detail=detail)
    db.session.add(log)


# ── 頁面路由 ────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('index.html')

@app.route('/health-mall')
def health_mall_page():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('health_mall.html')

@app.route('/analytics')
def analytics():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('analytics.html')

@app.route('/campaign-history')
def campaign_history_page():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('campaign_history.html')

@app.route('/baiwei')
def baiwei_page():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('baiwei.html')

@app.route('/campaign-match')
def campaign_match_page():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('campaign_match.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json()
        username = data.get('username', '')
        password = data.get('password', '')
        admin_password = os.environ.get('ADMIN_PASSWORD', '')
        user_password  = os.environ.get('USER_PASSWORD', '')
        if not admin_password or not user_password:
            return jsonify({'error': '系統設定錯誤，請聯絡管理員'}), 500
        if username == 'admin' and hmac.compare_digest(password, admin_password):
            session['user'] = 'admin'
            session['role'] = 'admin'
            return jsonify({'success': True})
        elif username == 'user' and hmac.compare_digest(password, user_password):
            session['user'] = 'user'
            session['role'] = 'user'
            return jsonify({'success': True})
        else:
            return jsonify({'error': '帳號或密碼錯誤'}), 401
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/audit-log')
def audit_log_page():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('audit_log.html')


# ── 診所管理 API ─────────────────────────────────────────────

@app.route('/api/clinics', methods=['GET'])
def get_clinics():
    """列出診所（分頁 + 資料庫層篩選）"""
    page        = max(1, request.args.get('page',     1,  type=int))
    per_page    = min(500, max(1, request.args.get('per_page', 50, type=int)))
    search      = request.args.get('search',      '').strip()
    city        = request.args.get('city',        '').strip()  # 縣市篩選
    specialty   = request.args.get('specialty',   '').strip()
    cooperation = request.args.get('cooperation', '').strip()  # 合作項目篩選
    sort_by     = request.args.get('sort',        '').strip()  # 排序欄位
    sort_dir    = request.args.get('dir',         'asc').strip().lower()  # asc / desc

    # 只顯示未軟刪除的診所
    query = Clinic.query.filter(Clinic.status != 'deleted')

    if search:
        # 搜尋診所名稱、電話或聯絡人
        query = query.filter(
            (Clinic.name.contains(search)) |
            (Clinic.phone.contains(search)) |
            (Clinic.contact_person.contains(search))
        )
    if city:
        query = query.filter(Clinic.region == city)
    if specialty:
        query = query.filter(Clinic.specialties.contains(specialty))
    if cooperation == 'yaodai':
        query = query.filter(Clinic.col_yaodai == True)
    elif cooperation == 'haibao':
        query = query.filter(Clinic.col_haibao == True)
    elif cooperation == 'paiyang':
        query = query.filter(Clinic.col_paiyang == True)
    elif cooperation == 'baiwei':
        query = query.filter(Clinic.col_baiwei == True)

    # 排序
    _sort_map = {
        'id': Clinic.id, 'region': Clinic.region, 'district': Clinic.district,
        'name': Clinic.name, 'specialties': Clinic.specialties,
        'phone': Clinic.phone, 'contact_person': Clinic.contact_person,
    }
    sort_col = _sort_map.get(sort_by, Clinic.id)
    order = sort_col.desc() if sort_dir == 'desc' else sort_col.asc()

    # 先算總筆數，再做分頁切片
    total       = query.count()
    total_pages = max(1, math.ceil(total / per_page))
    page        = min(page, total_pages)  # 防止超出頁數
    clinics     = query.order_by(order).offset((page - 1) * per_page).limit(per_page).all()

    def _fmt(c):
        return {
            'id':             c.id,
            'region':         c.region,
            'district':       c.district,
            'name':           c.name,
            'col_yaodai':     c.col_yaodai  or False,
            'col_haibao':     c.col_haibao  or False,
            'col_paiyang':    c.col_paiyang or False,
            'col_baiwei':     c.col_baiwei  or False,
            'specialties':    c.specialties,
            'address':        c.address,
            'phone':          c.phone,
            'contact_person': c.contact_person,
            'business_hours': c.business_hours,
            'note':           c.note,
            'created_at':     c.created_at.strftime('%Y-%m-%d %H:%M:%S') if c.created_at else None,
        }

    return jsonify({
        'clinics':     [_fmt(c) for c in clinics],
        'total':       total,
        'page':        page,
        'per_page':    per_page,
        'total_pages': total_pages,
    })


@app.route('/api/clinics/<int:clinic_id>', methods=['GET'])
def get_clinic(clinic_id):
    """取得單筆診所資料（供編輯 modal 使用）"""
    c = Clinic.query.get_or_404(clinic_id)
    return jsonify({
        'id':             c.id,
        'region':         c.region,
        'district':       c.district,
        'name':           c.name,
        'col_yaodai':     c.col_yaodai  or False,
        'col_haibao':     c.col_haibao  or False,
        'col_paiyang':    c.col_paiyang or False,
        'col_baiwei':     c.col_baiwei  or False,
        'specialties':    c.specialties,
        'address':        c.address,
        'phone':          c.phone,
        'contact_person': c.contact_person,
        'business_hours': c.business_hours,
        'note':           c.note,
        'created_at':     c.created_at.strftime('%Y-%m-%d %H:%M:%S') if c.created_at else None,
    })

@app.route('/api/clinics', methods=['POST'])
def create_clinic():
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    try:
        data = request.get_json()
        phone_val = data.get('phone')
        clinic = Clinic(
            region           = data.get('region'),
            district         = data.get('district'),
            name             = data.get('name'),
            col_yaodai       = data.get('col_yaodai', False),
            col_haibao       = data.get('col_haibao', False),
            col_paiyang      = data.get('col_paiyang', False),
            col_baiwei       = data.get('col_baiwei', False),
            specialties      = data.get('specialties'),
            address          = data.get('address'),
            phone            = phone_val,
            phone_normalized = _normalize_phone(phone_val) or None,  # 同步寫入正規化電話
            contact_person   = data.get('contact_person'),
            business_hours   = data.get('business_hours'),
            note             = data.get('note')
        )
        db.session.add(clinic)
        db.session.flush()  # 取得 clinic.id，尚未真正 commit
        _write_audit('新增', clinic.name, clinic.id)
        db.session.commit()
        return jsonify({'success': True, 'id': clinic.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'儲存失敗: {str(e)}'}), 500

@app.route('/api/clinics/<int:clinic_id>', methods=['PUT'])
def update_clinic(clinic_id):
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    try:
        clinic = Clinic.query.get_or_404(clinic_id)
        data = request.get_json()
        phone_val             = data.get('phone')
        clinic.region         = data.get('region')
        clinic.district       = data.get('district')
        clinic.name           = data.get('name')
        clinic.col_yaodai     = data.get('col_yaodai', False)
        clinic.col_haibao     = data.get('col_haibao', False)
        clinic.col_paiyang    = data.get('col_paiyang', False)
        clinic.col_baiwei     = data.get('col_baiwei', False)
        clinic.specialties    = data.get('specialties')
        clinic.address        = data.get('address')
        clinic.phone          = phone_val
        clinic.phone_normalized = _normalize_phone(phone_val) or None  # 同步更新正規化電話
        clinic.contact_person = data.get('contact_person')
        clinic.business_hours = data.get('business_hours')
        clinic.note           = data.get('note')
        _write_audit('編輯', clinic.name, clinic_id)
        db.session.commit()
        return jsonify({'success': True})
    except IntegrityError:
        db.session.rollback()
        return jsonify({'error': '此電話號碼已有其他診所使用，請確認是否為重複診所'}), 409
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'更新失敗: {str(e)}'}), 500

@app.route('/api/clinics/batch-edit', methods=['PUT'])
def batch_edit_clinics():
    """批次編輯多筆診所的縣市、科別、合作項目"""
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'error': '未選擇診所'}), 400

        fields = data.get('fields', {})
        if not fields:
            return jsonify({'error': '未指定修改內容'}), 400

        clinics = Clinic.query.filter(Clinic.id.in_(ids), Clinic.status != 'deleted').all()
        if not clinics:
            return jsonify({'error': '未找到符合的診所'}), 404

        updated = 0
        for clinic in clinics:
            if 'region' in fields and fields['region']:
                clinic.region = fields['region']
            if 'specialties' in fields and fields['specialties']:
                clinic.specialties = fields['specialties']
            if 'col_yaodai' in fields:
                clinic.col_yaodai = fields['col_yaodai']
            if 'col_haibao' in fields:
                clinic.col_haibao = fields['col_haibao']
            if 'col_paiyang' in fields:
                clinic.col_paiyang = fields['col_paiyang']
            if 'col_baiwei' in fields:
                clinic.col_baiwei = fields['col_baiwei']
            updated += 1

        _write_audit('批次編輯', f'共 {updated} 筆', 0, f'修改欄位: {list(fields.keys())}')
        db.session.commit()
        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'批次編輯失敗: {str(e)}'}), 500


@app.route('/api/clinics/<int:clinic_id>', methods=['DELETE'])
def delete_clinic(clinic_id):
    """軟刪除：只標記 status='deleted'，不實際刪除資料"""
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    clinic = Clinic.query.get_or_404(clinic_id)
    clinic.status     = 'deleted'
    clinic.deleted_at = datetime.utcnow()
    _write_audit('刪除', clinic.name, clinic_id)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/clinics/<int:clinic_id>/restore', methods=['PUT'])
def restore_clinic(clinic_id):
    """還原軟刪除的診所"""
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    clinic = Clinic.query.get_or_404(clinic_id)
    clinic.status     = 'active'
    clinic.deleted_at = None
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/clinics/<int:clinic_id>/permanent', methods=['DELETE'])
def permanent_delete_clinic(clinic_id):
    """永久刪除：先清除關聯資料再真正刪除"""
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    try:
        clinic = Clinic.query.get_or_404(clinic_id)
        clinic_name = clinic.name  # 先記錄名稱，因為刪除後無法取得
        CampaignClinic.query.filter_by(clinic_id=clinic_id).delete()
        BaiweiDoctor.query.filter_by(clinic_id=clinic_id).update({'clinic_id': None})
        HealthMall.query.filter_by(clinic_id=clinic_id).update({'clinic_id': None})
        db.session.delete(clinic)
        _write_audit('永久刪除', clinic_name, clinic_id)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'刪除失敗: {str(e)}'}), 500


@app.route('/api/clinics/deleted', methods=['GET'])
def get_deleted_clinics():
    """查詢所有已軟刪除的診所"""
    clinics = (Clinic.query
               .filter(Clinic.status == 'deleted')
               .order_by(Clinic.deleted_at.desc())
               .all())
    return jsonify([{
        'id':         c.id,
        'region':     c.region     or '',
        'district':   c.district   or '',
        'name':       c.name       or '',
        'phone':      c.phone      or '',
        'specialties': c.specialties or '',
        'deleted_at': c.deleted_at.strftime('%Y-%m-%d %H:%M') if c.deleted_at else '',
    } for c in clinics])

@app.route('/api/stats')
def get_stats():
    # 只計算未軟刪除的診所
    total = Clinic.query.filter(Clinic.status != 'deleted').count()
    return jsonify({'total': total})

@app.route('/api/audit-log')
def get_audit_log():
    """查詢操作記錄，支援以 action 過濾"""
    if 'user' not in session:
        return jsonify({'error': '未登入'}), 401
    action_filter = request.args.get('action', '').strip()
    query = AuditLog.query
    if action_filter:
        query = query.filter(AuditLog.action == action_filter)
    total_count = query.count()
    logs = query.order_by(AuditLog.created_at.desc()).limit(500).all()
    return jsonify({
        'total': total_count,
        'shown': len(logs),
        'logs': [{
            'id':          l.id,
            'action':      l.action      or '',
            'target_name': l.target_name or '',
            'target_id':   l.target_id,
            'detail':      l.detail      or '',
            'created_at':  l.created_at.strftime('%Y-%m-%d %H:%M:%S') if l.created_at else '',
        } for l in logs],
    })

@app.route('/api/clinics/duplicates')
def get_clinic_duplicates():
    """
    找出可能重複的診所：
    1. 診所名稱完全相同（SQL GROUP BY name）
    2. 電話號碼相同（SQL GROUP BY phone_normalized）
    3. 去除「診所/醫院/醫學中心/聯合/附設」後名稱相同（相似，Python）
    """
    from collections import defaultdict
    import re as _re
    from sqlalchemy import func

    def _brief(c):
        return {'id': c.id, 'name': c.name or '', 'phone': c.phone or '',
                'region': c.region or '', 'district': c.district or ''}

    # ── 1. 名稱完全相同 → SQL GROUP BY name ──────────────
    dup_name_rows = (
        db.session.query(Clinic.name, func.count(Clinic.id).label('cnt'))
        .filter(Clinic.status != 'deleted', Clinic.name != None)
        .group_by(Clinic.name)
        .having(func.count(Clinic.id) > 1)
        .all()
    )
    exact = []
    exact_name_set = set()
    for name, cnt in dup_name_rows:
        clinics = Clinic.query.filter(
            Clinic.status != 'deleted', Clinic.name == name
        ).order_by(Clinic.id).all()
        exact.append({'name': name, 'count': cnt, 'clinics': [_brief(c) for c in clinics]})
        exact_name_set.add(name)

    # ── 2. 電話重複 → SQL GROUP BY phone_normalized ──────
    dup_phone_rows = (
        db.session.query(Clinic.phone_normalized, func.count(Clinic.id).label('cnt'))
        .filter(Clinic.status != 'deleted',
                Clinic.phone_normalized != None,
                Clinic.phone_normalized != '')
        .group_by(Clinic.phone_normalized)
        .having(func.count(Clinic.id) > 1)
        .all()
    )
    for phone_norm, cnt in dup_phone_rows:
        clinics = Clinic.query.filter(
            Clinic.status != 'deleted', Clinic.phone_normalized == phone_norm
        ).order_by(Clinic.id).all()
        # 若所有重複診所名稱相同，已在 exact 組，跳過
        names = {c.name for c in clinics}
        if len(names) == 1 and list(names)[0] in exact_name_set:
            continue
        exact.append({
            'name':    f'電話重複：{clinics[0].phone or phone_norm}',
            'count':   cnt,
            'clinics': [_brief(c) for c in clinics],
        })

    # ── 3. 名稱相似（去除常見後綴詞後相同）→ Python ──────
    def simplify(name):
        s = _re.sub(r'診所|醫院|醫學中心|聯合|附設|小兒科|家醫科|耳鼻喉科|內科|外科|皮膚科|婦產科|中醫', '', name or '')
        return s.strip()

    # 只撈 _brief() 需要的 5 個欄位，不撈 SELECT *
    all_rows = (
        db.session.query(Clinic.id, Clinic.name, Clinic.phone, Clinic.region, Clinic.district)
        .filter(Clinic.status != 'deleted')
        .order_by(Clinic.id).all()
    )
    simplified_groups = defaultdict(list)
    for r in all_rows:
        if r.name:
            s = simplify(r.name)
            if s:
                simplified_groups[s].append(r)

    similar = []
    for simplified, rows in simplified_groups.items():
        if len(rows) > 1:
            names = {r.name for r in rows}
            if len(names) == 1 and list(names)[0] in exact_name_set:
                continue
            similar.append({
                'simplified': simplified,
                'count':      len(rows),
                'clinics':    [_brief(r) for r in rows],
            })

    return jsonify({'exact': exact, 'similar': similar})


@app.route('/api/clinics/merge', methods=['POST'])
def merge_clinics():
    """
    合併兩筆重複診所：
    - keep_id   保留的那筆
    - delete_id 刪除的那筆（活動/百位關聯轉移到 keep）
    """
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    data      = request.get_json() or {}
    keep_id   = data.get('keep_id')
    delete_id = data.get('delete_id')
    if not keep_id or not delete_id or keep_id == delete_id:
        return jsonify({'error': '參數錯誤'}), 400

    try:
        keep   = Clinic.query.get(keep_id)
        delete = Clinic.query.get(delete_id)
        if not keep:
            return jsonify({'error': f'找不到診所 id:{keep_id}'}), 404
        if not delete:
            return jsonify({'error': f'找不到診所 id:{delete_id}'}), 404

        # 用被刪筆補齊保留筆缺少的欄位
        for field in ('phone', 'phone_normalized', 'address', 'contact_person', 'specialties', 'region', 'district', 'note'):
            if not getattr(keep, field) and getattr(delete, field):
                setattr(keep, field, getattr(delete, field))

        # 合作項目取聯集
        for col in ('col_yaodai', 'col_haibao', 'col_paiyang', 'col_baiwei'):
            if getattr(delete, col):
                setattr(keep, col, True)

        # 轉移 campaign_clinic 關聯（跳過已存在的）
        for link in CampaignClinic.query.filter_by(clinic_id=delete_id).all():
            exists = CampaignClinic.query.filter_by(campaign_id=link.campaign_id, clinic_id=keep_id).first()
            if not exists:
                link.clinic_id = keep_id
            else:
                db.session.delete(link)

        # 轉移 baiwei_doctor 關聯
        BaiweiDoctor.query.filter_by(clinic_id=delete_id).update({'clinic_id': keep_id})

        # 軟刪除被合併的那筆
        delete.status = 'deleted'

        db.session.commit()

        _write_audit('合併', keep.name, keep_id,
                     detail=f'合併自 id:{delete_id}（{delete.name}）')
        return jsonify({'success': True})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'error': str(e)}), 500


@app.route('/api/specialties')
def get_specialties():
    """回傳系統標準科別白名單，供各頁面科別下拉選單使用"""
    return jsonify(STANDARD_SPECIALTIES)


@app.route('/api/clinics/search')
def search_clinics_for_hm():
    """供健康醫購新增時，從診所總表搜尋帶入資料"""
    q = request.args.get('q', '').strip()

    def _to_brief(c):
        return {
            'id':             c.id,
            'region':         c.region or '',
            'district':       c.district or '',
            'name':           c.name or '',
            'phone':          c.phone or '',
            'contact_person': c.contact_person or '',
            'specialties':    c.specialties or '',
            'address':        c.address or '',
        }

    if not q:
        clinics = Clinic.query.filter(Clinic.status != 'deleted').order_by(Clinic.name).limit(30).all()
        return jsonify([_to_brief(c) for c in clinics])

    q_digits = _normalize_phone(q)  # 純數字版搜尋字
    # 判斷是否像電話號碼：只含數字、空格、破折號，且至少 6 位數字
    is_phone_query = (
        bool(q_digits) and
        len(q_digits) >= 6 and
        bool(re.match(r'^[\d\s\-]+$', q))
    )

    if is_phone_query:
        # 電話搜尋：以 phone_normalized（純數字）做子字串比對，解決格式不一致問題
        # （DB 存 "02-2876-6955"，使用者搜 "0228766955" → 都轉純數字後比對）
        clinics = Clinic.query.filter(
            Clinic.status != 'deleted'
        ).filter(
            Clinic.name.contains(q) | Clinic.phone_normalized.contains(q_digits)
        ).order_by(Clinic.name).limit(30).all()
    else:
        clinics = Clinic.query.filter(
            Clinic.status != 'deleted'
        ).filter(
            Clinic.name.contains(q) | Clinic.phone.contains(q)
        ).order_by(Clinic.name).limit(30).all()

    return jsonify([_to_brief(c) for c in clinics])


# ── 健康醫購 API ─────────────────────────────────────────────

@app.route('/api/health-mall', methods=['GET'])
def get_health_mall():
    search    = request.args.get('search', '')
    region    = request.args.get('region', '')
    specialty = request.args.get('specialty', '')
    status    = request.args.get('status', '')

    query = HealthMall.query

    if search:
        query = query.filter(
            HealthMall.name.contains(search) |
            HealthMall.contact_person.contains(search)
        )
    if region:
        query = query.filter(HealthMall.region == region)
    if specialty:
        query = query.filter(HealthMall.specialties.contains(specialty))
    if status:
        query = query.filter(HealthMall.status == status)

    items = query.all()
    return jsonify([_hm_json(h) for h in items])

@app.route('/api/health-mall/stats')
def get_health_mall_stats():
    total  = HealthMall.query.count()
    active = HealthMall.query.filter(HealthMall.status == '合作中').count()
    return jsonify({'total': total, 'active': active})

@app.route('/api/health-mall', methods=['POST'])
def create_health_mall():
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    try:
        data = request.get_json()
        clinic_id_raw = data.get('clinic_id')
        hm = HealthMall(
            clinic_id      = int(clinic_id_raw) if clinic_id_raw else None,
            region         = data.get('region'),
            district       = data.get('district'),
            name           = data.get('name'),
            specialties    = data.get('specialties'),
            address        = data.get('address'),
            phone          = data.get('phone'),
            contact_person = data.get('contact_person'),
            status         = data.get('health_mall_status', '合作中'),
            start_date     = _parse_date(data.get('health_mall_start_date')),
            note           = data.get('health_mall_note'),
        )
        db.session.add(hm)
        db.session.commit()
        return jsonify({'success': True, 'id': hm.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'儲存失敗: {str(e)}'}), 500

@app.route('/api/health-mall/<int:hm_id>', methods=['PUT'])
def update_health_mall(hm_id):
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    try:
        hm = HealthMall.query.get_or_404(hm_id)
        data = request.get_json()
        clinic_id_raw = data.get('clinic_id')
        hm.clinic_id      = int(clinic_id_raw) if clinic_id_raw else hm.clinic_id
        hm.region         = data.get('region')
        hm.district       = data.get('district')
        hm.name           = data.get('name')
        hm.specialties    = data.get('specialties')
        hm.address        = data.get('address')
        hm.phone          = data.get('phone')
        hm.contact_person = data.get('contact_person')
        hm.status         = data.get('health_mall_status', '合作中')
        hm.start_date     = _parse_date(data.get('health_mall_start_date'))
        hm.note           = data.get('health_mall_note')
        hm.updated_at     = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'更新失敗: {str(e)}'}), 500

@app.route('/api/health-mall/<int:hm_id>', methods=['DELETE'])
def delete_health_mall(hm_id):
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    hm = HealthMall.query.get_or_404(hm_id)
    db.session.delete(hm)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/health-mall/export')
def export_health_mall():
    search    = request.args.get('search', '')
    region    = request.args.get('region', '')
    specialty = request.args.get('specialty', '')
    status    = request.args.get('status', '')

    query = HealthMall.query
    if search:
        query = query.filter(
            HealthMall.name.contains(search) |
            HealthMall.contact_person.contains(search)
        )
    if region:
        query = query.filter(HealthMall.region == region)
    if specialty:
        query = query.filter(HealthMall.specialties.contains(specialty))
    if status:
        query = query.filter(HealthMall.status == status)

    items = query.all()
    if not items:
        return jsonify({'error': '沒有符合條件的資料'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = '健康醫購'

    headers = ['縣市', '區域', '診所名稱', '科別', '地址', '電話', '負責人', '合作狀態', '開始日期', '備註']
    ws.append(headers)

    header_fill = PatternFill(start_color='11998E', end_color='11998E', fill_type='solid')
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for h in items:
        ws.append([
            h.region or '', h.district or '', h.name or '',
            h.specialties or '', h.address or '',
            h.phone or '', h.contact_person or '',
            h.status or '',
            h.start_date.strftime('%Y-%m-%d') if h.start_date else '',
            h.note or '',
        ])

    col_widths = [10, 10, 22, 18, 32, 15, 10, 10, 12, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'健康醫購_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ── 分析 API ─────────────────────────────────────────────────

@app.route('/api/analytics/stats')
def get_analytics_stats():
    """整合統計 API：診所 + 健康醫購（不含已軟刪除）"""
    # 診所總數（DB 計算，不全撈）
    total = Clinic.query.filter(Clinic.status != 'deleted').count()

    # 縣市分布（DB GROUP BY）
    region_rows = (
        db.session.query(Clinic.region, func.count(Clinic.id))
        .filter(Clinic.status != 'deleted', Clinic.region.isnot(None))
        .group_by(Clinic.region).all()
    )
    region_count = dict(region_rows)
    top_region = max(region_count.items(), key=lambda x: x[1]) if region_count else ('N/A', 0)

    # 科別分布（拆解組合科別，用 / 分割為個別科別計數）
    specialty_rows = (
        db.session.query(Clinic.specialties)
        .filter(Clinic.status != 'deleted', Clinic.specialties.isnot(None))
        .all()
    )
    specialty_list = []
    for (specialties,) in specialty_rows:
        if specialties:
            for s in re.split(r'[/,]', specialties):
                s = s.strip()
                if s:
                    specialty_list.append(s)
    specialty_count = Counter(specialty_list)
    # 按數量降序排列
    specialty_count = dict(specialty_count.most_common())

    # 合作項目分布（DB SUM/CASE，不全撈）
    col_row = db.session.query(
        func.sum(case((Clinic.col_yaodai  == True, 1), else_=0)).label('yaodai'),
        func.sum(case((Clinic.col_haibao  == True, 1), else_=0)).label('haibao'),
        func.sum(case((Clinic.col_paiyang == True, 1), else_=0)).label('paiyang'),
        func.sum(case((Clinic.col_baiwei  == True, 1), else_=0)).label('baiwei'),
    ).filter(Clinic.status != 'deleted').one()
    col_items = {
        '藥袋':     col_row.yaodai  or 0,
        '海報/立牌': col_row.haibao  or 0,
        '派樣':     col_row.paiyang or 0,
        '百位':     col_row.baiwei  or 0,
    }

    # 健康醫購（DB GROUP BY status）
    hm_total = HealthMall.query.count()
    hm_status_rows = (
        db.session.query(HealthMall.status, func.count(HealthMall.id))
        .group_by(HealthMall.status).all()
    )
    hm_status_map = dict(hm_status_rows)
    hm_active = hm_status_map.get('合作中', 0)
    hm_paused = hm_status_map.get('暫停',   0)
    hm_ended  = hm_status_map.get('結束',   0)
    hm_region_rows = (
        db.session.query(HealthMall.region, func.count(HealthMall.id))
        .filter(HealthMall.region.isnot(None))
        .group_by(HealthMall.region).all()
    )
    hm_region_count = dict(hm_region_rows)

    return jsonify({
        'total':            total,
        'hm_total':         hm_total,
        'hm_active':        hm_active,
        'top_region':       top_region[0],
        'top_region_count': top_region[1],
        'regions':      region_count,
        'specialties':  dict(specialty_count),
        'col_items':    col_items,
        'hm_status':    {'合作中': hm_active, '暫停': hm_paused, '結束': hm_ended},
        'hm_regions':   hm_region_count,
    })

@app.route('/api/analytics/regions')
def get_region_stats():
    rows = (
        db.session.query(Clinic.region, func.count(Clinic.id))
        .filter(Clinic.status != 'deleted', Clinic.region.isnot(None))
        .group_by(Clinic.region).all()
    )
    return jsonify({'regions': [r for r, _ in rows], 'counts': [c for _, c in rows]})

@app.route('/api/analytics/specialties')
def get_specialty_stats():
    rows = (
        db.session.query(Clinic.specialties)
        .filter(Clinic.status != 'deleted', Clinic.specialties.isnot(None))
        .all()
    )
    specialty_list = []
    for (specialties,) in rows:
        if specialties:
            for s in re.split(r'[/,]', specialties):
                s = s.strip()
                if s:
                    specialty_list.append(s)
    specialty_count = Counter(specialty_list)
    return jsonify({'specialties': list(specialty_count.keys()), 'counts': list(specialty_count.values())})

@app.route('/api/analytics/taiwan_map')
def get_taiwan_map_data():
    rows = (
        db.session.query(Clinic.region, func.count(Clinic.id))
        .filter(Clinic.status != 'deleted', Clinic.region.isnot(None))
        .group_by(Clinic.region).all()
    )
    return jsonify([{'name': r, 'value': v} for r, v in rows])


# ── 匯出/匯入 ────────────────────────────────────────────────

@app.route('/api/export', methods=['GET'])
def export_data():
    search      = request.args.get('search',      '').strip()
    city        = request.args.get('city',        '').strip()   # 新參數名
    specialty   = request.args.get('specialty',   '').strip()
    cooperation = request.args.get('cooperation', '').strip()   # 新參數名

    # 匯出只含未軟刪除的診所
    query = Clinic.query.filter(Clinic.status != 'deleted')
    if search:
        query = query.filter(
            Clinic.name.contains(search) |
            Clinic.phone.contains(search) |
            Clinic.contact_person.contains(search)
        )
    if city:
        query = query.filter(Clinic.region == city)
    if specialty:
        query = query.filter(Clinic.specialties.contains(specialty))
    if cooperation == 'yaodai':
        query = query.filter(Clinic.col_yaodai == True)
    elif cooperation == 'haibao':
        query = query.filter(Clinic.col_haibao == True)
    elif cooperation == 'paiyang':
        query = query.filter(Clinic.col_paiyang == True)
    elif cooperation == 'baiwei':
        query = query.filter(Clinic.col_baiwei == True)

    clinics = query.all()
    if not clinics:
        return jsonify({'error': '沒有符合條件的資料'}), 400

    output = export_clinics(clinics)
    filename = f'診所清單_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/import', methods=['POST'])
def import_data():
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    if 'file' not in request.files:
        return jsonify({'error': '沒有上傳檔案'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '沒有選擇檔案'}), 400
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '只接受 .xlsx 格式'}), 400

    filename = secure_filename(file.filename)
    temp_path = os.path.join('/tmp', filename)
    file.save(temp_path)
    dry_run = request.args.get('dry_run') == 'true'  # ?dry_run=true 時進入預覽模式
    with app.app_context():
        result = import_clinics(temp_path, db, Clinic, dry_run=dry_run)
    os.remove(temp_path)
    return jsonify(result)

@app.route('/api/import-custom', methods=['POST'])
def import_custom_data():
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    if 'file' not in request.files:
        return jsonify({'error': '沒有上傳檔案'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '沒有選擇檔案'}), 400
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '只接受 .xlsx 格式'}), 400

    filename = secure_filename(file.filename)
    temp_path = os.path.join('/tmp', filename)
    file.save(temp_path)
    dry_run = request.args.get('dry_run') == 'true'  # ?dry_run=true 時進入預覽模式
    with app.app_context():
        result = import_custom_clinics(temp_path, db, Clinic, dry_run=dry_run)
    os.remove(temp_path)
    return jsonify(result)

@app.route('/api/health-mall/import', methods=['POST'])
def import_health_mall_data():
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    if 'file' not in request.files:
        return jsonify({'error': '沒有上傳檔案'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '沒有選擇檔案'}), 400
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '只接受 .xlsx 格式'}), 400

    filename = secure_filename(file.filename)
    temp_path = os.path.join('/tmp', filename)
    file.save(temp_path)
    dry_run = request.args.get('dry_run') == 'true'  # ?dry_run=true 時進入預覽模式
    with app.app_context():
        result = import_health_mall(temp_path, db, HealthMall, dry_run=dry_run)
    os.remove(temp_path)
    return jsonify(result)


# ── 活動比對 API ─────────────────────────────────────────────

@app.route('/api/campaign/match', methods=['POST'])
def campaign_match():
    if 'file' not in request.files:
        return jsonify({'error': '沒有上傳檔案'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '只接受 .xlsx 格式'}), 400

    filename = secure_filename(file.filename)
    temp_path = os.path.join('/tmp', filename)
    file.save(temp_path)

    try:
        wb = load_workbook(temp_path)
        ws = wb.active
        os.remove(temp_path)

        # 自動偵測 header 行：掃描前 5 列，找到含「電話」的那行
        # （與 import_campaign_clinics 邏輯一致，支援第一列為標題列的 Excel）
        header_row_idx = None
        header = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            stripped = [str(v).strip() if v is not None else '' for v in row]
            if '電話' in stripped:
                header_row_idx = row_idx
                header = stripped
                break

        if header_row_idx is None:
            return jsonify({'error': 'Excel 缺少必要欄位：電話（掃描前 5 列均未找到）'}), 400

        col = {name: idx for idx, name in enumerate(header)}

        def _get(row, name):
            idx = col.get(name)
            if idx is None or idx >= len(row):
                return ''
            return str(row[idx]).strip() if row[idx] is not None else ''

        uploaded = []
        consecutive_blank = 0  # 連續空白行計數器
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not any(row):
                consecutive_blank += 1
                if consecutive_blank >= 5:  # 連續5筆空白即停止
                    break
                continue
            consecutive_blank = 0  # 有資料就重置
            phone_raw  = _get(row, '電話')
            region_val = _get(row, '縣市')
            phone_fmt  = format_phone(phone_raw, region_val or None)  # 先補區碼
            uploaded.append({
                'region':   region_val,
                'district': _get(row, '區域'),
                'name':     _get(row, '診所名稱'),
                'phone':    phone_raw,
                'phone_n':  _normalize_phone(phone_fmt),              # 再 normalize → 完整含區碼數字
            })

        # 只撈比對需要的欄位，減少資料傳輸量
        clinic_rows = (
            db.session.query(
                Clinic.id, Clinic.region, Clinic.district, Clinic.name,
                Clinic.specialties, Clinic.address, Clinic.phone,
                Clinic.contact_person, Clinic.phone_normalized
            )
            .filter(Clinic.status != 'deleted')
            .all()
        )
        def _row_brief(r):
            return {
                'id': r.id, 'region': r.region or '', 'district': r.district or '',
                'name': r.name or '', 'specialties': r.specialties or '',
                'address': r.address or '', 'phone': r.phone or '',
                'contact_person': r.contact_person or '',
            }
        phone_to_row = {r.phone_normalized: r for r in clinic_rows if r.phone_normalized}

        uploaded_phones = {u['phone_n'] for u in uploaded if u['phone_n']}

        matched = []
        for u in uploaded:
            if u['phone_n'] and u['phone_n'] in phone_to_row:
                matched.append(_row_brief(phone_to_row[u['phone_n']]))

        not_joined = []
        for r in clinic_rows:
            if not r.phone_normalized or r.phone_normalized not in uploaded_phones:
                not_joined.append(_row_brief(r))

        not_in_system = []
        for u in uploaded:
            if not u['phone_n'] or u['phone_n'] not in phone_to_row:
                not_in_system.append({
                    'region':   u['region'],
                    'district': u['district'],
                    'name':     u['name'],
                    'phone':    u['phone'],
                })

        result = {
            'matched':       matched,
            'not_joined':    not_joined,
            'not_in_system': not_in_system,
        }

        # 儲存比對歷史
        campaign_name = request.form.get('campaign_name', file.filename or '未命名')
        import json as _json
        history = MatchHistory(
            campaign_name=campaign_name,
            matched_count=len(matched),
            not_joined_count=len(not_joined),
            not_in_system_count=len(not_in_system),
            result_json=_json.dumps(result, ensure_ascii=False),
        )
        db.session.add(history)
        db.session.commit()
        result['history_id'] = history.id

        return jsonify(result)

    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({'error': f'比對失敗: {str(e)}'}), 500


@app.route('/api/campaign/match-history', methods=['GET'])
def list_match_history():
    """列出比對歷史記錄（最新在前）"""
    rows = MatchHistory.query.order_by(MatchHistory.created_at.desc()).limit(50).all()
    return jsonify([{
        'id': r.id,
        'campaign_name': r.campaign_name,
        'matched_count': r.matched_count,
        'not_joined_count': r.not_joined_count,
        'not_in_system_count': r.not_in_system_count,
        'created_at': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
    } for r in rows])


@app.route('/api/campaign/match-history/<int:history_id>', methods=['GET'])
def get_match_history(history_id):
    """取得單筆比對歷史的完整結果"""
    import json as _json
    row = MatchHistory.query.get_or_404(history_id)
    result = _json.loads(row.result_json) if row.result_json else {}
    result['campaign_name'] = row.campaign_name
    result['created_at'] = row.created_at.strftime('%Y-%m-%d %H:%M') if row.created_at else ''
    return jsonify(result)


@app.route('/api/campaign/match-history/<int:history_id>', methods=['DELETE'])
def delete_match_history(history_id):
    """刪除單筆比對歷史"""
    row = MatchHistory.query.get_or_404(history_id)
    db.session.delete(row)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/campaign/export-not-joined')
def export_not_joined():
    ids_str  = request.args.get('ids', '')
    campaign = request.args.get('campaign', '活動')

    if not ids_str:
        return jsonify({'error': '沒有診所 ID'}), 400

    try:
        ids = [int(i) for i in ids_str.split(',') if i.strip().isdigit()]
    except ValueError:
        return jsonify({'error': '無效的 ID 格式'}), 400

    clinics = Clinic.query.filter(Clinic.id.in_(ids), Clinic.status != 'deleted').all()
    if not clinics:
        return jsonify({'error': '沒有符合條件的資料'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = '未參加診所'
    ws.append(['縣市', '區域', '診所名稱', '科別', '地址', '電話', '負責人', '備註'])
    for c in clinics:
        ws.append([
            c.region or '', c.district or '', c.name or '',
            c.specialties or '', c.address or '',
            c.phone or '', c.contact_person or '', c.note or '',
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'{campaign}_未參加診所_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ── 活動記錄 API ─────────────────────────────────────────────

@app.route('/api/campaigns', methods=['GET'])
def get_campaigns():
    campaigns = Campaign.query.order_by(Campaign.year.desc(), Campaign.created_at.desc()).all()
    # 一條 SQL 取得所有活動的診所數量，避免 N+1
    counts = dict(
        db.session.query(CampaignClinic.campaign_id, func.count(CampaignClinic.id))
        .group_by(CampaignClinic.campaign_id)
        .all()
    )
    result = []
    for c in campaigns:
        result.append({
            'id':                c.id,
            'name':              c.name,
            'brand':             c.brand or '',
            'year':              c.year,
            'month':             c.month,
            'note':              c.note or '',
            'cooperation_items': c.cooperation_items or '',
            'cooperation_other': c.cooperation_other or '',
            'clinic_count':      counts.get(c.id, 0),
            'created_at':        c.created_at.strftime('%Y-%m-%d') if c.created_at else '',
        })
    return jsonify(result)

@app.route('/api/campaigns', methods=['POST'])
def create_campaign():
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    data              = request.get_json()
    brand             = (data.get('brand') or '').strip()
    year              = data.get('year')
    month             = data.get('month')
    note              = data.get('note') or ''
    cooperation_items = (data.get('cooperation_items') or '').strip()
    cooperation_other = (data.get('cooperation_other') or '').strip()
    name  = (data.get('name') or '').strip() or f"{brand}{year or ''}".strip()
    if not name:
        return jsonify({'error': '請輸入活動名稱或品牌'}), 400
    c = Campaign(name=name, brand=brand or None,
                 year=int(year) if year else None,
                 month=int(month) if month else None,
                 note=note or None,
                 cooperation_items=cooperation_items or None,
                 cooperation_other=cooperation_other or None)
    db.session.add(c)
    db.session.commit()
    return jsonify({'success': True, 'id': c.id, 'name': c.name})

@app.route('/api/campaigns/<int:campaign_id>', methods=['DELETE'])
def delete_campaign(campaign_id):
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    campaign = Campaign.query.get_or_404(campaign_id)
    CampaignClinic.query.filter_by(campaign_id=campaign_id).delete()
    db.session.delete(campaign)
    db.session.commit()
    return jsonify({'success': True})

# 編輯活動資訊（品牌、年份、月份、備註）
@app.route('/api/campaigns/<int:campaign_id>', methods=['PUT'])
def update_campaign(campaign_id):
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    campaign = Campaign.query.get_or_404(campaign_id)
    data              = request.get_json()
    brand             = (data.get('brand') or '').strip()
    year              = data.get('year')
    month             = data.get('month')
    note              = (data.get('note') or '').strip()
    cooperation_items = (data.get('cooperation_items') or '').strip()
    cooperation_other = (data.get('cooperation_other') or '').strip()
    # 年份和月份為必填
    if not year:
        return jsonify({'error': '年份為必填'}), 400
    if not month:
        return jsonify({'error': '月份為必填'}), 400
    campaign.brand             = brand or None
    campaign.year              = int(year)
    campaign.month             = int(month)
    campaign.note              = note or None
    campaign.cooperation_items = cooperation_items or None
    campaign.cooperation_other = cooperation_other or None
    db.session.commit()
    return jsonify({'success': True, 'id': campaign.id, 'name': campaign.name})

@app.route('/api/campaigns/<int:campaign_id>/clinics', methods=['GET'])
def get_campaign_clinics(campaign_id):
    Campaign.query.get_or_404(campaign_id)
    region    = request.args.get('region', '')
    specialty = request.args.get('specialty', '')
    links = (CampaignClinic.query
             .options(joinedload(CampaignClinic.clinic))
             .filter_by(campaign_id=campaign_id)
             .all())
    result = []
    for link in links:
        c = link.clinic
        if not c:
            continue
        if region and c.region != region:
            continue
        if specialty and (not c.specialties or specialty not in c.specialties):
            continue
        result.append({
            'id':             c.id,
            'region':         c.region or '',
            'district':       c.district or '',
            'name':           c.name or '',
            'specialties':    c.specialties or '',
            'address':        c.address or '',
            'phone':          c.phone or '',
            'contact_person': c.contact_person or '',
            'note':           c.note or '',
            'joined_at':      link.joined_at.strftime('%Y-%m-%d') if link.joined_at else '',
        })
    return jsonify(result)

@app.route('/api/campaigns/<int:campaign_id>/clinics/<int:clinic_id>', methods=['PUT'])
def update_campaign_clinic(campaign_id, clinic_id):
    """更新活動中某間診所的基本資料（直接修改 clinic table）"""
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    # 確認診所確實屬於此活動
    link = CampaignClinic.query.filter_by(campaign_id=campaign_id, clinic_id=clinic_id).first()
    if not link:
        return jsonify({'error': '此診所不在該活動中'}), 404
    clinic = Clinic.query.get_or_404(clinic_id)
    data = request.get_json() or {}
    # 只更新有傳入的欄位，避免意外清空
    if 'region'         in data: clinic.region         = data['region']
    if 'district'       in data: clinic.district       = data['district']
    if 'name'           in data: clinic.name           = data['name']
    if 'specialties'    in data: clinic.specialties    = data['specialties']
    if 'phone'          in data:
        clinic.phone            = data['phone']
        clinic.phone_normalized = _normalize_phone(data['phone']) or None  # 同步更新正規化電話
    if 'contact_person' in data: clinic.contact_person = data['contact_person']
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/campaigns/<int:campaign_id>/import', methods=['POST'])
def import_campaign_clinics(campaign_id):
    print(f"[DEBUG] campaign_import called, campaign_id={campaign_id}, dry_run={request.args.get('dry_run')}, files={list(request.files.keys())}", flush=True)
    if session.get('role') != 'admin':
        print(f"[DEBUG] returning 403 権限不足", flush=True)
        return jsonify({'error': '權限不足'}), 403
    campaign = Campaign.query.get_or_404(campaign_id)
    if 'file' not in request.files:
        print(f"[DEBUG] returning 400 沒有上傳檔案", flush=True)
        return jsonify({'error': '沒有上傳檔案'}), 400
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        print(f"[DEBUG] returning 400 非xlsx, filename={file.filename}", flush=True)
        return jsonify({'error': '只接受 .xlsx 格式'}), 400

    filename  = secure_filename(file.filename)
    temp_path = os.path.join('/tmp', filename)

    try:
        file.save(temp_path)
        print(f"[DEBUG] file saved to {temp_path}", flush=True)
        wb = load_workbook(temp_path, read_only=True, data_only=True)
        print(f"[DEBUG] workbook loaded (read_only)", flush=True)
        ws = wb.active
        os.remove(temp_path)

        # 前三行 debug 資訊 + 自動偵測 header 行：合併一次掃描，減少記憶體
        preview_rows = []
        header_row_idx = None
        header = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            row = list(row)
            if row_idx <= 3:
                preview_rows.append([str(v) if v is not None else '' for v in row])
            if header_row_idx is None:
                stripped = [str(v).strip() if v is not None else '' for v in row]
                if '電話' in stripped:
                    header_row_idx = row_idx
                    header = stripped

        if header_row_idx is None:
            wb.close()
            print(f"[DEBUG] returning 400 找不到標題行, preview={preview_rows}", flush=True)
            return jsonify({
                'error': '找不到含「電話」的標題行（掃描前 5 行）',
                'debug_preview': preview_rows,
            }), 400

        col = {name: idx for idx, name in enumerate(header)}

        # 支援雙名稱欄位
        def _resolve(row, *candidates):
            for field in candidates:
                idx = col.get(field)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
            return ''

        # 建立電話 → 診所物件對應表（用 phone_normalized 為鍵，不含軟刪除）
        phone_to_clinic = {}
        for c in Clinic.query.filter(Clinic.status != 'deleted').all():
            if c.phone_normalized:
                phone_to_clinic[c.phone_normalized] = c

        in_campaign = {cc.clinic_id for cc in
                       CampaignClinic.query.filter_by(campaign_id=campaign_id).all()}

        dry_run = request.args.get('dry_run') == 'true'  # ?dry_run=true 時進入預覽模式
        new_clinics_count = updated_count = recorded = already_in_campaign = 0
        errors = []
        processed_clinics = []  # 記錄本次處理過的診所，匯入後連動合作項目
        preview_create = []  # 新增預覽（最多 5 筆）
        preview_update = []  # 更新預覽（最多 5 筆）

        data_start_row = header_row_idx + 1
        consecutive_blank = 0  # 連續空白行計數器
        for row_num, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            if not any(row):
                consecutive_blank += 1
                if consecutive_blank >= 5:  # 連續5筆空白即停止
                    break
                continue
            consecutive_blank = 0  # 有資料就重置
            phone_raw  = _resolve(row, '電話')
            phone_fmt  = format_phone(phone_raw, _resolve(row, '縣市') or None)  # 先補區碼
            normalized = _normalize_phone(phone_fmt)                              # 再正規化（含區碼的完整數字）
            if not normalized:
                errors.append(f'第{row_num}列：缺少電話')
                continue

            if normalized in phone_to_clinic:
                # 電話已存在 → 更新診所基本資料（phone / phone_normalized 不動）
                clinic = phone_to_clinic[normalized]
                new_region  = _resolve(row, '縣市') or None
                new_district = _resolve(row, '區域') or None
                new_name_raw = _resolve(row, '診所名稱', '院名')
                new_spec    = normalize_specialty(_resolve(row, '科別')) or None
                new_addr    = _resolve(row, '地址', '院址') or None
                new_contact = _resolve(row, '負責人', '聯絡人') or None
                if new_region:   clinic.region         = new_region
                if new_district: clinic.district       = new_district
                if new_name_raw:
                    from import_custom import _parse_name
                    new_name, _ = _parse_name(new_name_raw)
                    if new_name: clinic.name = new_name
                if new_spec:    clinic.specialties    = new_spec
                if new_addr:    clinic.address        = new_addr
                if new_contact: clinic.contact_person = new_contact
                updated_count += 1
                clinic_name_for_preview = _resolve(row, '診所名稱', '院名')
                if len(preview_update) < 5:
                    preview_update.append({'name': clinic_name_for_preview, 'phone': phone_fmt, 'action': 'update', 'region': _resolve(row, '縣市')})
            else:
                raw_name = _resolve(row, '診所名稱', '院名')
                if not raw_name:
                    errors.append(f'第{row_num}列：總表無此電話且缺少診所名稱')
                    continue
                # 套用院名清理邏輯（括號/斜線內容移到備註）
                from import_custom import _parse_name
                name, extracted_note = _parse_name(raw_name)
                if not name:
                    errors.append(f'第{row_num}列：清理後診所名稱為空（原始值：{raw_name}）')
                    continue
                clinic = Clinic(
                    region           = _resolve(row, '縣市') or None,
                    district         = _resolve(row, '區域') or None,
                    name             = name,
                    specialties      = normalize_specialty(_resolve(row, '科別')) or None,
                    address          = _resolve(row, '地址', '院址') or None,
                    phone            = phone_fmt,
                    phone_normalized = normalized or None,  # format_phone 後的完整數字
                    contact_person   = _resolve(row, '負責人', '聯絡人') or None,
                    note             = extracted_note or None,
                )
                try:
                    # 用 savepoint 隔離 INSERT，UniqueViolation 只回滾該筆
                    sp = db.session.begin_nested()
                    db.session.add(clinic)
                    db.session.flush()
                    phone_to_clinic[normalized] = clinic
                    new_clinics_count += 1
                    if len(preview_create) < 5:
                        preview_create.append({'name': name, 'phone': phone_fmt, 'action': 'create', 'region': _resolve(row, '縣市')})
                except IntegrityError:
                    # 並發情況：INSERT 失敗但 phone_to_clinic 未命中，改為更新
                    sp.rollback()
                    clinic = Clinic.query.filter_by(phone_normalized=normalized).first()
                    if not clinic:
                        errors.append(f'第{row_num}列：電話衝突但查無既有診所，跳過')
                        continue
                    phone_to_clinic[normalized] = clinic
                    updated_count += 1
                    if len(preview_update) < 5:
                        preview_update.append({'name': name, 'phone': phone_fmt, 'action': 'update', 'region': _resolve(row, '縣市')})

            processed_clinics.append(clinic)

            # 不管診所是新增還是已存在，都立即同步合作項目（只設 True，不取消已有勾選）
            # 放在 continue 之前，確保 already_in_campaign 的診所也會更新
            coop_items = campaign.cooperation_items or ''
            if coop_items:
                if '藥袋' in coop_items:
                    clinic.col_yaodai = True
                if '海報' in coop_items or '立牌' in coop_items:
                    clinic.col_haibao = True
                if '派樣' in coop_items:
                    clinic.col_paiyang = True
                if '百位' in coop_items:
                    clinic.col_baiwei = True

            if clinic.id in in_campaign:
                already_in_campaign += 1
                continue
            # dry_run 模式下不加入 CampaignClinic，僅計數
            if not dry_run:
                db.session.add(CampaignClinic(campaign_id=campaign_id, clinic_id=clinic.id))
            in_campaign.add(clinic.id)
            recorded += 1

        # dry_run 模式：rollback 確保不寫入，回傳預覽結果
        if dry_run:
            print(f"[DEBUG] dry_run returning: would_create={new_clinics_count}, would_update={updated_count}, would_skip={already_in_campaign}, errors={len(errors)}, preview={len(preview_create+preview_update)}", flush=True)
            wb.close()
            try:
                db.session.rollback()
            except Exception:
                pass
            print(f"[DEBUG] dry_run rollback done, sending jsonify response", flush=True)
            return jsonify({
                'dry_run':             True,
                'would_create':        new_clinics_count,
                'would_update':        updated_count,
                'would_skip':          already_in_campaign,
                'errors':              errors,
                'preview':             preview_create + preview_update,
            })

        print(f"[DEBUG] commit returning: new={new_clinics_count}, updated={updated_count}, recorded={recorded}, skip={already_in_campaign}, errors={len(errors)}", flush=True)
        wb.close()
        db.session.commit()
        return jsonify({
            'success':             True,
            'new_clinics':         new_clinics_count,
            'updated':             updated_count,
            'recorded':            recorded,
            'already_in_campaign': already_in_campaign,
            'errors':              errors,
            'debug_header_row':    header_row_idx,
            'debug_columns':       header,
            'debug_preview':       preview_rows,
        })

    except Exception as e:
        import traceback
        print(f"[DEBUG] exception caught: {type(e).__name__}: {e}", flush=True)
        print(f"[DEBUG] traceback: {traceback.format_exc()}", flush=True)
        try:
            wb.close()
        except Exception:
            pass
        if os.path.exists(temp_path):
            os.remove(temp_path)
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'error': f'匯入失敗: {str(e)}'}), 500

@app.route('/api/campaigns/<int:campaign_id>/clinics/all', methods=['DELETE'])
def clear_campaign_clinics(campaign_id):
    """清空指定活動的所有診所名單（只刪 campaign_clinic 關聯，不動 clinic 資料）"""
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    Campaign.query.get_or_404(campaign_id)
    deleted = CampaignClinic.query.filter_by(campaign_id=campaign_id).delete()
    db.session.commit()
    return jsonify({'success': True, 'deleted': deleted})


@app.route('/api/campaigns/<int:campaign_id>/export', methods=['GET'])
def export_campaign_clinics(campaign_id):
    campaign = Campaign.query.get_or_404(campaign_id)
    links    = CampaignClinic.query.filter_by(campaign_id=campaign_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = campaign.name[:28]
    headers  = ['縣市', '區域', '診所名稱', '科別', '地址', '電話', '聯絡人', '備註']
    ws.append(headers)
    fill = PatternFill(start_color='E96C2C', end_color='E96C2C', fill_type='solid')
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = fill
        cell.alignment = Alignment(horizontal='center')
    for link in links:
        c = link.clinic
        if not c:
            continue
        ws.append([c.region or '', c.district or '', c.name or '',
                   c.specialties or '', c.address or '',
                   c.phone or '', c.contact_person or '', c.note or ''])
    for i, w in enumerate([10, 10, 22, 18, 32, 15, 10, 24], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'{campaign.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@app.route('/api/clinics/<int:clinic_id>/campaigns', methods=['GET'])
def get_clinic_campaigns(clinic_id):
    links = CampaignClinic.query.filter_by(clinic_id=clinic_id).all()
    result = []
    for link in links:
        c = link.campaign
        if c.year and c.month:
            campaign_date = f'{c.year}年{c.month}月'
        elif c.year:
            campaign_date = f'{c.year}年'
        else:
            campaign_date = ''
        result.append({
            'campaign_id':   c.id,
            'campaign_name': c.name,
            'brand':         c.brand or '',
            'year':          c.year,
            'campaign_date': campaign_date,
        })
    return jsonify(result)


# ── 百位醫師 API ──────────────────────────────────────────────

@app.route('/api/baiwei', methods=['GET'])
def get_baiwei():
    specialty = request.args.get('specialty', '')
    query = BaiweiDoctor.query
    if specialty:
        # 用 LIKE 比對，支援「家醫科/外科」這類複合科別欄位
        query = query.filter(BaiweiDoctor.specialty.like(f'%{specialty}%'))
    items = query.order_by(BaiweiDoctor.region, BaiweiDoctor.district, BaiweiDoctor.clinic_name).all()
    return jsonify([{
        'id':          i.id,
        'clinic_id':   i.clinic_id,
        'clinic_name': i.clinic_name or '',
        'region':      i.region or '',
        'district':    i.district or '',
        'address':     i.address or '',
        'phone':       i.phone or '',
        'doctor_name': i.doctor_name or '',
        'specialty':   i.specialty or '',
    } for i in items])


@app.route('/api/baiwei/stats', methods=['GET'])
def get_baiwei_stats():
    total = BaiweiDoctor.query.count()
    return jsonify({'total': total})


@app.route('/api/baiwei/specialties', methods=['GET'])
def get_baiwei_specialties():
    """
    從 baiwei_doctor 的 specialty 欄位動態產生科別選單。
    以斜線拆分複合科別（如「家醫科/外科」→「家醫科」、「外科」），
    去重後排序，並排除指定科別。
    """
    EXCLUDE = {'泌尿科', '牙科', '產後護理之家', '醫美'}
    rows = db.session.query(BaiweiDoctor.specialty).filter(BaiweiDoctor.specialty.isnot(None)).all()
    seen = set()
    for (val,) in rows:
        # 同時支援斜線、頓號、逗號三種分隔符號
        for part in re.split(r'[/、,]', val):
            part = part.strip()
            if part and part not in EXCLUDE:
                seen.add(part)
    return jsonify(sorted(seen))


@app.route('/api/baiwei/import', methods=['POST'])
def import_baiwei():
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    if 'file' not in request.files:
        return jsonify({'error': '沒有上傳檔案'}), 400
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '只接受 .xlsx 格式'}), 400

    filename  = secure_filename(file.filename)
    temp_path = os.path.join('/tmp', filename)
    file.save(temp_path)

    try:
        wb = load_workbook(temp_path)
        ws = wb.active
        os.remove(temp_path)

        # 偵測 header 行（找含「電話」的那行）
        header_row_idx = None
        header = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            stripped = [str(v).strip() if v is not None else '' for v in row]
            if '電話' in stripped:
                header_row_idx = row_idx
                header = stripped
                break
        if header_row_idx is None:
            return jsonify({'error': '找不到含「電話」的標題行（掃描前 5 行）'}), 400

        col = {name: idx for idx, name in enumerate(header)}

        def _get(row, *keys):
            for k in keys:
                idx = col.get(k)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
            return ''

        # 建立電話 → clinic 的快查表（不含軟刪除）
        phone_to_clinic = {}
        for c in Clinic.query.filter(Clinic.status != 'deleted').all():
            np = _normalize_phone(c.phone)
            if np:
                phone_to_clinic[np] = c

        # 建立已存在的百位醫師組合（phone + doctor_name）防重複匯入
        existing_keys = {
            (_normalize_phone(d.phone), (d.doctor_name or '').strip())
            for d in BaiweiDoctor.query.all()
        }

        dry_run = request.args.get('dry_run') == 'true'  # ?dry_run=true 時進入預覽模式
        new_count = already_count = 0
        errors = []
        preview = []  # 預覽清單（最多 10 筆）

        consecutive_blank = 0  # 連續空白行計數器
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            if not any(row):
                consecutive_blank += 1
                if consecutive_blank >= 5:  # 連續5筆空白即停止
                    break
                continue
            consecutive_blank = 0  # 有資料就重置

            phone_raw   = _get(row, '電話')
            doctor_name = _get(row, '醫師', '醫師名字')
            phone_fmt_bw = format_phone(phone_raw, _get(row, '縣市') or None)  # 先補區碼
            normalized   = _normalize_phone(phone_fmt_bw)                       # 再正規化（含區碼的完整數字）

            if not normalized:
                errors.append(f'第{row_num}列：缺少電話')
                continue
            if not doctor_name:
                errors.append(f'第{row_num}列：缺少醫師名字')
                continue

            # 重複檢查
            if (normalized, doctor_name) in existing_keys:
                already_count += 1
                if len(preview) < 10:
                    preview.append({'name': _get(row, '診所名稱', '院名'), 'phone': phone_fmt_bw, 'action': 'skip', 'region': _get(row, '縣市')})
                continue

            # 比對或新增診所
            clinic = phone_to_clinic.get(normalized)
            if clinic is None:
                raw_name = _get(row, '診所名稱', '院名')
                if not raw_name:
                    errors.append(f'第{row_num}列：總表無此電話且缺少診所名稱')
                    continue
                from import_custom import _parse_name
                name, extracted_note = _parse_name(raw_name)
                if not name:
                    errors.append(f'第{row_num}列：診所名稱清理後為空（原始：{raw_name}）')
                    continue
                clinic = Clinic(
                    region           = _get(row, '縣市') or None,
                    district         = _get(row, '區域') or None,
                    name             = name,
                    address          = _get(row, '地址') or None,
                    phone            = phone_fmt_bw,
                    phone_normalized = normalized or None,  # format_phone 後的完整數字
                    note             = extracted_note or None,
                )
                try:
                    # 用 savepoint 隔離 INSERT，UniqueViolation 只回滾該筆
                    sp = db.session.begin_nested()
                    db.session.add(clinic)
                    db.session.flush()
                    phone_to_clinic[normalized] = clinic
                except IntegrityError:
                    sp.rollback()
                    # 資料庫已有此電話（並發或舊資料），查出現有診所繼續使用
                    clinic = Clinic.query.filter_by(phone_normalized=normalized).first()
                    if not clinic:
                        errors.append(f'第{row_num}列：電話衝突但查無既有診所，跳過')
                        continue
                    phone_to_clinic[normalized] = clinic

            # 設定 col_baiwei
            clinic.col_baiwei = True

            doc = BaiweiDoctor(
                clinic_id   =clinic.id,
                clinic_name =_get(row, '診所名稱', '院名') or clinic.name,
                region      =_get(row, '縣市') or clinic.region,
                district    =_get(row, '區域') or clinic.district,
                address     =_get(row, '地址') or clinic.address,
                phone       =phone_fmt_bw,
                doctor_name =doctor_name,
                specialty   = normalize_specialty(_get(row, '科別')) or None,
            )
            db.session.add(doc)
            existing_keys.add((normalized, doctor_name))
            new_count += 1
            if len(preview) < 10:
                preview.append({'name': _get(row, '診所名稱', '院名') or (clinic.name if clinic else ''), 'phone': phone_fmt_bw, 'action': 'create', 'region': _get(row, '縣市')})

        # dry_run 模式：rollback 確保不寫入，回傳預覽結果
        if dry_run:
            db.session.rollback()
            return jsonify({
                'dry_run':      True,
                'would_create': new_count,
                'would_update': 0,
                'would_skip':   already_count,
                'errors':       errors,
                'preview':      preview,
            })

        db.session.commit()
        return jsonify({'success': True, 'new': new_count, 'already_exists': already_count, 'errors': errors})

    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        db.session.rollback()
        return jsonify({'error': f'匯入失敗: {str(e)}'}), 500


@app.route('/api/baiwei/duplicates', methods=['GET'])
def get_baiwei_duplicates():
    """回傳百位總表中重複的醫師群組（同電話＋同醫師姓名）"""
    all_docs = BaiweiDoctor.query.order_by(BaiweiDoctor.id).all()
    groups = {}
    for d in all_docs:
        key = (_normalize_phone(d.phone), (d.doctor_name or '').strip())
        if key not in groups:
            groups[key] = []
        groups[key].append({
            'id':          d.id,
            'doctor_name': d.doctor_name or '',
            'clinic_name': d.clinic_name or '',
            'region':      d.region or '',
            'specialty':   d.specialty or '',
            'phone':       d.phone or '',
        })
    duplicates = [v for v in groups.values() if len(v) > 1]
    return jsonify({'groups': duplicates, 'total_extra': sum(len(g) - 1 for g in duplicates)})


@app.route('/api/baiwei/deduplicate', methods=['POST'])
def deduplicate_baiwei():
    """刪除百位總表重複記錄，每組保留 id 最小的那筆"""
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    all_docs = BaiweiDoctor.query.order_by(BaiweiDoctor.id).all()
    groups = {}
    for d in all_docs:
        key = (_normalize_phone(d.phone), (d.doctor_name or '').strip())
        if key not in groups:
            groups[key] = []
        groups[key].append(d)
    deleted = 0
    for docs in groups.values():
        if len(docs) > 1:
            for dup in docs[1:]:  # 保留第一筆（id最小），刪除其餘
                BaiweiParticipation.query.filter_by(doctor_id=dup.id).delete()
                db.session.delete(dup)
                deleted += 1
    db.session.commit()
    return jsonify({'success': True, 'deleted': deleted})


@app.route('/api/baiwei/clear-all', methods=['POST'])
def clear_all_baiwei():
    """清空百位醫師總表（不動診所管理總表）"""
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    BaiweiParticipation.query.delete()
    count = BaiweiDoctor.query.count()
    BaiweiDoctor.query.delete()
    # 將所有診所的 col_baiwei 重設為 False
    Clinic.query.filter_by(col_baiwei=True).update({'col_baiwei': False})
    db.session.commit()
    return jsonify({'success': True, 'deleted': count})


@app.route('/api/baiwei/<int:doc_id>', methods=['DELETE'])
def delete_baiwei(doc_id):
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    doc = BaiweiDoctor.query.get_or_404(doc_id)
    clinic_id = doc.clinic_id
    db.session.delete(doc)
    db.session.flush()
    # 若該診所已無其他百位醫師記錄，將 col_baiwei 設回 False
    if clinic_id:
        remaining = BaiweiDoctor.query.filter_by(clinic_id=clinic_id).count()
        if remaining == 0:
            clinic = Clinic.query.get(clinic_id)
            if clinic:
                clinic.col_baiwei = False
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/baiwei/export', methods=['GET'])
def export_baiwei():
    specialty = request.args.get('specialty', '')
    query = BaiweiDoctor.query
    if specialty:
        # 與列表 API 一致，用 LIKE 比對複合科別欄位（如「家醫科/外科」）
        query = query.filter(BaiweiDoctor.specialty.like(f'%{specialty}%'))
    items = query.order_by(BaiweiDoctor.region, BaiweiDoctor.district, BaiweiDoctor.clinic_name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '百位醫師'
    headers = ['縣市', '區域', '診所名稱', '醫師名字', '科別', '電話']
    ws.append(headers)

    fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')

    for d in items:
        ws.append([d.region or '', d.district or '', d.clinic_name or '',
                   d.doctor_name or '', d.specialty or '', d.phone or ''])

    for i, w in enumerate([10, 10, 24, 14, 16, 14], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'百位醫師_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# ── 百位活動 API ──────────────────────────────────────────────

@app.route('/api/baiwei-campaigns', methods=['GET'])
def get_baiwei_campaigns():
    campaigns = BaiweiCampaign.query.order_by(BaiweiCampaign.year.desc(), BaiweiCampaign.id.desc()).all()
    return jsonify([{
        'id':    c.id,
        'name':  c.name,
        'year':  c.year,
        'month': c.month,
        'note':  c.note or '',
        'count': BaiweiParticipation.query.filter_by(campaign_id=c.id).count(),
    } for c in campaigns])


@app.route('/api/baiwei-campaigns', methods=['POST'])
def create_baiwei_campaign():
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': '活動名稱不可空白'}), 400
    c = BaiweiCampaign(name=name, year=data.get('year') or None, month=data.get('month') or None, note=data.get('note') or None)
    db.session.add(c)
    db.session.commit()
    return jsonify({'id': c.id, 'name': c.name, 'year': c.year, 'month': c.month, 'note': c.note or '', 'count': 0})


@app.route('/api/baiwei-campaigns/<int:campaign_id>', methods=['DELETE'])
def delete_baiwei_campaign(campaign_id):
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    c = BaiweiCampaign.query.get_or_404(campaign_id)
    BaiweiParticipation.query.filter_by(campaign_id=campaign_id).delete()
    db.session.delete(c)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/baiwei-campaigns/<int:campaign_id>/doctors', methods=['GET'])
def get_baiwei_campaign_doctors(campaign_id):
    BaiweiCampaign.query.get_or_404(campaign_id)
    specialty = request.args.get('specialty', '')
    q = (db.session.query(BaiweiParticipation, BaiweiDoctor)
         .join(BaiweiDoctor, BaiweiParticipation.doctor_id == BaiweiDoctor.id)
         .filter(BaiweiParticipation.campaign_id == campaign_id))
    if specialty:
        q = q.filter(BaiweiDoctor.specialty.like(f'%{specialty}%'))
    rows = q.order_by(BaiweiDoctor.region, BaiweiDoctor.clinic_name).all()
    return jsonify([{
        'participation_id': p.id,
        'id':          d.id,
        'doctor_name': d.doctor_name or '',
        'clinic_name': d.clinic_name or '',
        'region':      d.region or '',
        'district':    d.district or '',
        'specialty':   d.specialty or '',
        'phone':       d.phone or '',
    } for p, d in rows])


@app.route('/api/baiwei-campaigns/<int:campaign_id>/doctors', methods=['POST'])
def add_baiwei_campaign_doctors(campaign_id):
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    BaiweiCampaign.query.get_or_404(campaign_id)
    data = request.get_json() or {}
    doctor_ids = data.get('doctor_ids', [])
    if not doctor_ids:
        return jsonify({'error': '未指定醫師'}), 400
    added = 0
    for did in doctor_ids:
        if not BaiweiParticipation.query.filter_by(doctor_id=did, campaign_id=campaign_id).first():
            db.session.add(BaiweiParticipation(doctor_id=did, campaign_id=campaign_id))
            added += 1
    db.session.commit()
    return jsonify({'success': True, 'added': added})


@app.route('/api/baiwei-campaigns/<int:campaign_id>/doctors/<int:doctor_id>', methods=['DELETE'])
def remove_baiwei_campaign_doctor(campaign_id, doctor_id):
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    p = BaiweiParticipation.query.filter_by(campaign_id=campaign_id, doctor_id=doctor_id).first_or_404()
    db.session.delete(p)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/baiwei-campaigns/<int:campaign_id>/export', methods=['GET'])
def export_baiwei_campaign(campaign_id):
    campaign = BaiweiCampaign.query.get_or_404(campaign_id)
    rows = (db.session.query(BaiweiParticipation, BaiweiDoctor)
            .join(BaiweiDoctor, BaiweiParticipation.doctor_id == BaiweiDoctor.id)
            .filter(BaiweiParticipation.campaign_id == campaign_id)
            .order_by(BaiweiDoctor.region, BaiweiDoctor.clinic_name).all())
    wb = Workbook()
    ws = wb.active
    ws.title = campaign.name[:28]
    headers = ['縣市', '區域', '診所名稱', '醫師名字', '科別', '電話']
    ws.append(headers)
    fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
    for p, d in rows:
        ws.append([d.region or '', d.district or '', d.clinic_name or '',
                   d.doctor_name or '', d.specialty or '', d.phone or ''])
    for i, w in enumerate([10, 10, 22, 12, 15, 15], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'{campaign.name}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/api/baiwei', methods=['POST'])
def add_baiwei_doctor():
    """手動新增單筆百位醫師"""
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    data = request.get_json() or {}
    doctor_name = (data.get('doctor_name') or '').strip()
    phone_raw   = (data.get('phone') or '').strip()
    if not doctor_name:
        return jsonify({'error': '醫師名字不可空白'}), 400
    if not phone_raw:
        return jsonify({'error': '電話不可空白'}), 400
    phone_fmt  = format_phone(phone_raw, data.get('region') or None)
    normalized = _normalize_phone(phone_fmt)
    # 比對診所
    clinic = Clinic.query.filter(Clinic.status != 'deleted').filter_by(phone_normalized=normalized).first()
    if clinic is None and data.get('clinic_name'):
        from import_custom import _parse_name
        name, extracted_note = _parse_name(data.get('clinic_name'))
        clinic = Clinic(
            region=data.get('region') or None, district=data.get('district') or None,
            name=name, phone=phone_fmt, phone_normalized=normalized or None,
        )
        db.session.add(clinic)
        db.session.flush()
    if clinic:
        clinic.col_baiwei = True
    doc = BaiweiDoctor(
        clinic_id   = clinic.id if clinic else None,
        clinic_name = (data.get('clinic_name') or '').strip() or (clinic.name if clinic else None),
        region      = data.get('region') or (clinic.region if clinic else None),
        district    = data.get('district') or (clinic.district if clinic else None),
        phone       = phone_fmt,
        doctor_name = doctor_name,
        specialty   = normalize_specialty(data.get('specialty') or '') or None,
    )
    db.session.add(doc)
    db.session.commit()
    return jsonify({'success': True, 'id': doc.id})


@app.route('/api/baiwei-campaigns/<int:campaign_id>/import', methods=['POST'])
def import_baiwei_campaign(campaign_id):
    """匯入 Excel 到指定百位活動（同時更新醫師主表）"""
    if session.get('role') != 'admin':
        return jsonify({'error': '權限不足'}), 403
    campaign = BaiweiCampaign.query.get_or_404(campaign_id)
    if 'file' not in request.files:
        return jsonify({'error': '沒有上傳檔案'}), 400
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '只接受 .xlsx 格式'}), 400

    filename  = secure_filename(file.filename)
    temp_path = os.path.join('/tmp', filename)
    file.save(temp_path)

    try:
        wb = load_workbook(temp_path, read_only=True, data_only=True)
        ws = wb.active
        os.remove(temp_path)

        # 偵測 header 行
        header_row_idx = None
        header = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            stripped = [str(v).strip() if v is not None else '' for v in row]
            if '電話' in stripped:
                header_row_idx = row_idx
                header = stripped
                break
        if header_row_idx is None:
            return jsonify({'error': '找不到含「電話」的標題行（掃描前 5 行）'}), 400

        col = {name: idx for idx, name in enumerate(header)}

        def _get(row, *keys):
            for k in keys:
                idx = col.get(k)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
            return ''

        phone_to_clinic = {}
        for c in Clinic.query.filter(Clinic.status != 'deleted').all():
            np = _normalize_phone(c.phone)
            if np:
                phone_to_clinic[np] = c

        existing_keys = {
            (_normalize_phone(d.phone), (d.doctor_name or '').strip())
            for d in BaiweiDoctor.query.all()
        }
        # 已在此活動的醫師 id set
        in_campaign = {p.doctor_id for p in BaiweiParticipation.query.filter_by(campaign_id=campaign_id).all()}

        dry_run = request.args.get('dry_run') == 'true'
        new_count = already_count = added_count = 0
        errors = []
        preview = []

        consecutive_blank = 0
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            if not any(row):
                consecutive_blank += 1
                if consecutive_blank >= 5:
                    break
                continue
            consecutive_blank = 0

            phone_raw   = _get(row, '電話')
            doctor_name = _get(row, '醫師', '醫師名字')
            phone_fmt   = format_phone(phone_raw, _get(row, '縣市') or None)
            normalized  = _normalize_phone(phone_fmt)

            if not normalized:
                errors.append(f'第{row_num}列：缺少電話')
                continue
            if not doctor_name:
                errors.append(f'第{row_num}列：缺少醫師名字')
                continue

            # 找或建 clinic
            clinic = phone_to_clinic.get(normalized)
            if clinic is None:
                raw_name = _get(row, '診所名稱', '院名')
                if not raw_name:
                    errors.append(f'第{row_num}列：總表無此電話且缺少診所名稱')
                    continue
                from import_custom import _parse_name
                name, extracted_note = _parse_name(raw_name)
                if not name:
                    errors.append(f'第{row_num}列：診所名稱清理後為空')
                    continue
                clinic = Clinic(
                    region=_get(row, '縣市') or None, district=_get(row, '區域') or None,
                    name=name, address=_get(row, '地址') or None,
                    phone=phone_fmt, phone_normalized=normalized or None,
                    note=extracted_note or None,
                )
                try:
                    sp = db.session.begin_nested()
                    db.session.add(clinic)
                    db.session.flush()
                    phone_to_clinic[normalized] = clinic
                except IntegrityError:
                    sp.rollback()
                    clinic = Clinic.query.filter_by(phone_normalized=normalized).first()
                    if not clinic:
                        errors.append(f'第{row_num}列：電話衝突但查無既有診所，跳過')
                        continue
                    phone_to_clinic[normalized] = clinic

            clinic.col_baiwei = True

            # 找或建 baiwei_doctor
            key = (normalized, doctor_name)
            if key in existing_keys:
                # 醫師已在主表，只加活動關聯
                doc = BaiweiDoctor.query.filter_by(phone=phone_fmt, doctor_name=doctor_name).first()
                already_count += 1
            else:
                doc = BaiweiDoctor(
                    clinic_id=clinic.id, clinic_name=_get(row, '診所名稱', '院名') or clinic.name,
                    region=_get(row, '縣市') or clinic.region, district=_get(row, '區域') or clinic.district,
                    address=_get(row, '地址') or clinic.address, phone=phone_fmt, doctor_name=doctor_name,
                    specialty=normalize_specialty(_get(row, '科別')) or None,
                )
                db.session.add(doc)
                db.session.flush()
                existing_keys.add(key)
                new_count += 1

            # 加活動關聯
            if doc and doc.id not in in_campaign:
                db.session.add(BaiweiParticipation(doctor_id=doc.id, campaign_id=campaign_id))
                in_campaign.add(doc.id)
                added_count += 1

            if len(preview) < 10:
                preview.append({'name': _get(row, '診所名稱', '院名') or (clinic.name if clinic else ''),
                                 'phone': phone_fmt, 'action': 'create' if key not in existing_keys else 'skip',
                                 'region': _get(row, '縣市')})

        if dry_run:
            db.session.rollback()
            return jsonify({'dry_run': True, 'would_create': new_count, 'would_update': already_count,
                            'would_add': added_count, 'errors': errors, 'preview': preview})

        db.session.commit()
        return jsonify({'success': True, 'new': new_count, 'already_exists': already_count,
                        'added': added_count, 'errors': errors})

    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        db.session.rollback()
        return jsonify({'error': f'匯入失敗: {str(e)}'}), 500


# ── 工具函式 ─────────────────────────────────────────────────

def _normalize_phone(phone):
    return re.sub(r'\D', '', str(phone)) if phone else ''

def _clinic_brief(c):
    return {
        'id':             c.id,
        'region':         c.region or '',
        'district':       c.district or '',
        'name':           c.name or '',
        'specialties':    c.specialties or '',
        'address':        c.address or '',
        'phone':          c.phone or '',
        'contact_person': c.contact_person or '',
    }

def _parse_date(s):
    if s:
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except ValueError:
            return None
    return None

def _hm_json(h):
    """HealthMall 物件 → JSON dict（保持與前端相容的欄位名稱）"""
    return {
        'id':                   h.id,
        'clinic_id':            h.clinic_id,
        'region':               h.region or '',
        'district':             h.district or '',
        'name':                 h.name or '',
        'specialties':          h.specialties or '',
        'address':              h.address or '',
        'phone':                h.phone or '',
        'contact_person':       h.contact_person or '',
        'health_mall_status':   h.status or '合作中',
        'health_mall_start_date': h.start_date.strftime('%Y-%m-%d') if h.start_date else '',
        'health_mall_note':     h.note or '',
    }


with app.app_context():
    db.create_all()
    # 逐一補上歷次新增的欄位，IF NOT EXISTS 確保重複執行也安全
    _migrations = [
        'ALTER TABLE campaign ADD COLUMN IF NOT EXISTS month INTEGER',
        'ALTER TABLE campaign ADD COLUMN IF NOT EXISTS cooperation_items VARCHAR(200)',
        'ALTER TABLE campaign ADD COLUMN IF NOT EXISTS cooperation_other VARCHAR(200)',
        # 新增正規化電話欄位（移除非數字後的電話），用於唯一性約束
        'ALTER TABLE clinic ADD COLUMN IF NOT EXISTS phone_normalized VARCHAR(50)',
        # 填入現有診所的正規化電話
        "UPDATE clinic SET phone_normalized = regexp_replace(phone, '[^0-9]', '', 'g') WHERE phone IS NOT NULL AND phone_normalized IS NULL",
        # 建立唯一索引（排除空值，避免無電話的診所互相衝突）
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_clinic_phone_normalized ON clinic(phone_normalized) WHERE phone_normalized IS NOT NULL AND phone_normalized <> ''",
        """CREATE TABLE IF NOT EXISTS baiwei_doctor (
            id SERIAL PRIMARY KEY,
            clinic_id INTEGER REFERENCES clinic(id) ON DELETE SET NULL,
            clinic_name VARCHAR(200),
            region VARCHAR(50),
            district VARCHAR(50),
            address VARCHAR(300),
            phone VARCHAR(50),
            doctor_name VARCHAR(100),
            specialty VARCHAR(200),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        # 效能索引：常用篩選欄位
        'CREATE INDEX IF NOT EXISTS idx_clinic_status ON clinic(status)',
        'CREATE INDEX IF NOT EXISTS idx_clinic_region ON clinic(region)',
        'CREATE INDEX IF NOT EXISTS idx_clinic_status_region ON clinic(status, region)',
        'CREATE INDEX IF NOT EXISTS idx_audit_log_created_at ON audit_log(created_at DESC)',
        'CREATE INDEX IF NOT EXISTS idx_audit_log_action ON audit_log(action)',
        'CREATE INDEX IF NOT EXISTS idx_campaign_clinic_campaign_id ON campaign_clinic(campaign_id)',
        'CREATE INDEX IF NOT EXISTS idx_campaign_clinic_clinic_id ON campaign_clinic(clinic_id)',
        # 移除廢棄欄位（已被 col_yaodai/col_haibao/col_paiyang/col_baiwei 取代）
        'ALTER TABLE clinic DROP COLUMN IF EXISTS media_items',
        'ALTER TABLE baiwei_campaign ADD COLUMN IF NOT EXISTS month INTEGER',
        # 百位活動表
        """CREATE TABLE IF NOT EXISTS baiwei_campaign (
            id SERIAL PRIMARY KEY,
            name VARCHAR(200) NOT NULL,
            year INTEGER,
            note TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        # 百位活動參加記錄（多對多）
        """CREATE TABLE IF NOT EXISTS baiwei_participation (
            id SERIAL PRIMARY KEY,
            doctor_id INTEGER NOT NULL REFERENCES baiwei_doctor(id) ON DELETE CASCADE,
            campaign_id INTEGER NOT NULL REFERENCES baiwei_campaign(id) ON DELETE CASCADE,
            CONSTRAINT uq_baiwei_participation UNIQUE (doctor_id, campaign_id)
        )""",
        # 比對歷史記錄表
        """CREATE TABLE IF NOT EXISTS match_history (
            id SERIAL PRIMARY KEY,
            campaign_name VARCHAR(200),
            matched_count INTEGER DEFAULT 0,
            not_joined_count INTEGER DEFAULT 0,
            not_in_system_count INTEGER DEFAULT 0,
            result_json TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
    ]
    for _sql in _migrations:
        try:
            with db.engine.connect() as _conn:
                _conn.execute(text(_sql))
                _conn.commit()
        except Exception:
            pass

    # 一次性修正：補上 baiwei_doctor 電話區碼（phone 不以 '0' 開頭的記錄）
    try:
        _docs_to_fix = BaiweiDoctor.query.filter(
            BaiweiDoctor.phone.isnot(None),
            ~BaiweiDoctor.phone.like('0%')
        ).all()
        for _d in _docs_to_fix:
            _new_phone = format_phone(_d.phone, _d.region or None)
            if _new_phone != _d.phone:
                _d.phone = _new_phone
        if _docs_to_fix:
            db.session.commit()
    except Exception:
        db.session.rollback()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8081, debug=True)
