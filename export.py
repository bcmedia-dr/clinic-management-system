from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io

def export_clinics(clinics):
    """匯出診所資料到 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "診所清單"
    
    # 設定欄位標題（按照你要的順序）
    headers = ['縣市', '區域', '診所名稱', '科別', '地址', '電話', '負責人']
    
    # 寫入標題列（第1行）
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 寫入資料（從第2行開始）
    for row_num, clinic in enumerate(clinics, 2):
        row_data = [
            clinic.region or '',
            clinic.district or '',
            clinic.name or '',
            clinic.specialties or '',
            clinic.address or '',
            clinic.phone or '',
            clinic.contact_person or ''
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # 調整欄寬
    column_widths = {
        'A': 12,  # 縣市
        'B': 12,  # 區域
        'C': 25,  # 診所名稱
        'D': 20,  # 科別
        'E': 35,  # 地址
        'F': 15,  # 電話
        'G': 12   # 負責人
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 設定行高
    ws.row_dimensions[1].height = 25
    
    # 儲存到記憶體
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
